#!/usr/bin/env python3
"""料金モデル v2 (全員同一料金・10社優遇ゼロ・成果報酬 Phase 3+) を反映する。

合意内容 (2026-04-12):
  - 見学会予約費: ¥50,000/件 (税別、全員同じ) — Phase 1 から
  - 撮影プラン: ¥150,000/本 (全員同じ) — Phase 2 から
  - 月額掲載 ベーシック: ¥30,000/月 (全員同じ) — Phase 3 から
  - 月額掲載 プロ: ¥80,000/月 — Phase 3 から
  - 成果報酬 3%: 上限なし、無料相談 / AI診断 / 会員登録済みの成約のみ — Phase 3 から
  - SaaS ライト: ¥50,000/月 — Phase 4 から
  - SaaS プロ: ¥150,000/月 — Phase 4 から
  - 既存10社優遇: AI診断結果での優先送客のみ (料金優遇ゼロ)

このスクリプトは以下の xlsx を再生成する:
  - docs/最新版/02_管理シート/MVP実行計画.xlsx
  - docs/最新版/02_管理シート/フェーズ別収益モデル.xlsx
  - docs/最新版/02_管理シート/動線設計シート.xlsx
  - docs/最新版/04_経理関係/請求・P&L管理.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_MGMT = ROOT / "docs" / "最新版" / "02_管理シート"
OUT_KEIRI = ROOT / "docs" / "最新版" / "04_経理関係"
OUT_MGMT.mkdir(parents=True, exist_ok=True)
OUT_KEIRI.mkdir(parents=True, exist_ok=True)

# ─── Styling ──────────────────────────────────────────────────────

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ORANGE = PatternFill("solid", fgColor="E8740C")
CREAM = PatternFill("solid", fgColor="FFF8F0")
DARK = PatternFill("solid", fgColor="3D2200")
GRAY = PatternFill("solid", fgColor="F3F4F6")
GREEN = PatternFill("solid", fgColor="D1FAE5")
YELLOW = PatternFill("solid", fgColor="FEF3C7")
RED = PatternFill("solid", fgColor="FEE2E2")
BLUE = PatternFill("solid", fgColor="DBEAFE")

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True, size=10)
REG = Font(size=10)
SMALL = Font(size=9, color="6B7280")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def H(ws, row: int, headers: list[str], widths: list[int]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = DARK
        c.font = WHITE_BOLD
        c.alignment = CENTER
        c.border = BORDER
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


def R(ws, row: int, values: list, fill=None, align=LEFT, bold=False) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(bold=bold, size=10) if bold else REG
        c.alignment = align
        c.border = BORDER
        if fill:
            c.fill = fill
    ws.row_dimensions[row].height = 22


def T(ws, row: int, title: str, cols: int, subtitle: str | None = None) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=18, color="3D2200")
    c.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 34
    row += 1
    if subtitle:
        c = ws.cell(row=row, column=1, value=subtitle)
        c.font = SMALL
        c.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        row += 1
    return row + 1


# ─── 保守的な月次想定 (全員同一料金ベース) ────────────────────────────

# 月次 KPI (見学会件数・撮影本数・月額契約社数・成果報酬件数・SaaS社数)
# Phase 3 以降に成果報酬と月額掲載が加算される
MONTHLY_PROJECTION = [
    # (month_label, phase, booking, filming, subscription, success_cases, saas_basic, saas_pro)
    ("5月", "P1", 10, 0, 0, 0, 0, 0),
    ("6月", "P1", 15, 0, 0, 0, 0, 0),
    ("7月", "P2", 20, 1, 0, 0, 0, 0),
    ("8月", "P2", 25, 2, 0, 0, 0, 0),
    ("9月", "P3", 30, 3, 3, 1, 0, 0),
    ("10月", "P3", 35, 3, 5, 2, 0, 0),
    ("11月", "P3", 40, 4, 8, 2, 0, 0),
    ("12月", "P4", 45, 4, 10, 3, 2, 0),
    ("1月", "P4", 50, 5, 12, 3, 4, 0),
    ("2月", "P4", 55, 5, 13, 4, 6, 1),
    ("3月", "P5", 60, 6, 15, 4, 8, 1),
    ("4月", "P5", 65, 6, 17, 5, 10, 2),
]

# 料金マスター
PRICE_BOOKING = 50_000       # ¥50,000/件
PRICE_FILMING = 150_000      # ¥150,000/本
PRICE_SUB_BASIC = 30_000     # ¥30,000/月
PRICE_SUB_PRO = 80_000       # ¥80,000/月
PRICE_SAAS_LITE = 50_000     # ¥50,000/月
PRICE_SAAS_PRO = 150_000     # ¥150,000/月
RATE_SUCCESS = 0.03          # 3%
ASSUMED_CONTRACT = 30_000_000  # ¥3,000万 想定
SUCCESS_UNIT = int(ASSUMED_CONTRACT * RATE_SUCCESS)  # ¥900,000


def calc_monthly_revenue(item) -> dict:
    month, phase, booking_n, filming_n, sub_n, success_n, saas_lite_n, saas_pro_n = item
    booking_rev = booking_n * PRICE_BOOKING
    filming_rev = filming_n * PRICE_FILMING
    sub_rev = sub_n * PRICE_SUB_BASIC  # ベーシック想定
    success_rev = success_n * SUCCESS_UNIT
    saas_rev = saas_lite_n * PRICE_SAAS_LITE + saas_pro_n * PRICE_SAAS_PRO
    total = booking_rev + filming_rev + sub_rev + success_rev + saas_rev
    return {
        "month": month,
        "phase": phase,
        "booking_n": booking_n,
        "filming_n": filming_n,
        "sub_n": sub_n,
        "success_n": success_n,
        "saas_lite_n": saas_lite_n,
        "saas_pro_n": saas_pro_n,
        "booking_rev": booking_rev,
        "filming_rev": filming_rev,
        "sub_rev": sub_rev,
        "success_rev": success_rev,
        "saas_rev": saas_rev,
        "total": total,
    }


PROJECTIONS = [calc_monthly_revenue(item) for item in MONTHLY_PROJECTION]


# ══════════════════════════════════════════════════════════════════
# ① MVP実行計画.xlsx を再生成
# ══════════════════════════════════════════════════════════════════

def build_execution_plan() -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "目次"
    row = T(
        ws, 1, "ぺいほーむ MVP 実行計画", 3,
        "2026-05-01 リリースに向けた実行管理シート。全シートは「誰が・いつまでに・何をするか」で書かれています。",
    )
    H(ws, row, ["#", "シート名", "用途"], [6, 32, 60])
    row += 1
    for code, name, use in [
        ("①", "リリース前ToDo (4月)", "4/11〜4/30 の作業タスク一覧。担当と期日で管理"),
        ("②", "Phase1 週次ToDo", "5/1〜6/30 の週次運用作業。毎週月曜の朝会で使う"),
        ("③", "Phase別機能追加スケジュール", "Phase 1〜5 で何を、いつ、誰が復活させるか"),
        ("④", "工務店10社 対応管理", "10社それぞれの連絡・契約・撮影・送客の進捗管理"),
        ("⑤", "リリース前チェックリスト", "4/30 までに完了するべき項目の一覧"),
        ("⑥", "月次KPIトラッキング", "リリース後12ヶ月のKPI目標と実績記録欄 (全員同一料金)"),
        ("⑦", "非公開パス復活スケジュール", "40+ 非公開パスの復活タイミング一覧"),
    ]:
        R(ws, row, [code, name, use])
        row += 1

    # ── ① リリース前ToDo ──
    ws = wb.create_sheet("① リリース前ToDo (4月)")
    row = T(ws, 1, "① リリース前ToDo (4月)", 7, "2026-05-01 リリースに向けた確定タスク")
    H(ws, row, ["#", "タスク", "担当", "期日", "所要", "ステータス", "備考"], [5, 50, 14, 12, 8, 14, 30])
    row += 1
    for t in [
        ("4-1", "Vercel 本番プロジェクト作成", "エンジニア", "2026-04-20", "1h", "未着手", "独自ドメイン payhome.jp を紐付け"),
        ("4-2", "本番用 .env 設定 (NextAuth, Supabase)", "エンジニア", "2026-04-20", "2h", "未着手", "NEXTAUTH_SECRET は新規生成"),
        ("4-3", "Supabase 本番マイグレーション適用", "エンジニア", "2026-04-21", "2h", "未着手", "users/leads/events/contact_preferences テーブル"),
        ("4-4", "本番DBに初期データ投入 (工務店10社)", "エンジニア", "2026-04-22", "3h", "未着手", "builders-data.ts と同期"),
        ("4-5", "GA4 計測タグ埋め込み動作確認", "エンジニア", "2026-04-23", "1h", "未着手", "lead_source を追跡可能に"),
        ("4-6", "Cloudflare Turnstile 有効化", "エンジニア", "2026-04-23", "2h", "未着手", "signup/contact/event booking"),
        ("4-7", "Sentry エラー監視セットアップ", "エンジニア", "2026-04-24", "2h", "未着手", "Vercel 統合も検討"),
        ("4-8", "17公開ページの E2E スモークテスト", "エンジニア", "2026-04-25", "3h", "未着手", "Playwright で自動化"),
        ("4-9", "40+ 非公開ページの 404 検証", "エンジニア", "2026-04-25", "1h", "未着手", "スクリプトで一括実行"),
        ("4-10", "Search Console 登録 + sitemap 送信", "PM", "2026-04-26", "1h", "未着手", "payhome.jp/sitemap.xml"),
        ("4-11", "工務店10社へリリース予告メール送信", "営業", "2026-04-20", "2h", "未着手", "料金プラン_MVP版.docx 添付 (全員同一料金)"),
        ("4-12", "工務店10社と料金プラン最終合意 (¥50,000/件)", "営業", "2026-04-27", "−", "未着手", "5/1 以前に書面締結"),
        ("4-13", "営業FAQ の準備・全スタッフ共有", "営業", "2026-04-24", "2h", "未着手", "想定Q&A 20問、10社優遇=AI診断優先送客のみ"),
        ("4-14", "問合せ対応フロー・担当割り当て", "PM", "2026-04-25", "1h", "未着手", "24h以内返信ルール"),
        ("4-15", "経理: 請求書テンプレート作成", "経理", "2026-04-23", "1h", "未着手", "見学会予約 ¥50,000 税別・振込先"),
        ("4-16", "経理: 入金確認フロー設定", "経理", "2026-04-25", "1h", "未着手", "freee で管理"),
        ("4-17", "プレスリリース原稿作成", "代表", "2026-04-27", "2h", "未着手", "5/1 AM 9:00 配信想定"),
        ("4-18", "YouTube リリース告知動画 収録・公開", "代表", "2026-04-29", "4h", "未着手", "5/1 AM 6:00 公開"),
        ("4-19", "SNS リリース告知予約投稿", "代表", "2026-04-30", "1h", "未着手", "X/Instagram/Threads"),
        ("4-20", "本番切替リハーサル", "エンジニア", "2026-04-29", "2h", "未着手", "Vercel preview → production"),
        ("4-21", "5/1 AM 6:00 本番リリース", "エンジニア", "2026-05-01", "1h", "未着手", "代表と営業もスタンバイ"),
        ("4-22", "リリース後の初日モニタリング", "全員", "2026-05-01", "8h", "未着手", "AM 6:00〜PM 22:00 障害対応体制"),
    ]:
        R(ws, row, list(t))
        row += 1
    ws.freeze_panes = "A3"

    # ── ② Phase1 週次ToDo ──
    ws = wb.create_sheet("② Phase1 週次ToDo")
    row = T(ws, 1, "② Phase 1 週次運用ToDo (5/1〜6/30)", 6, "毎週月曜 9:00 の朝会で確認")
    H(ws, row, ["曜日", "時刻", "タスク", "担当", "頻度", "判断基準"], [8, 10, 45, 14, 10, 40])
    row += 1
    for item in [
        ("月", "9:00", "前週の見学会予約・リード・診断数の集計", "PM", "毎週", "先週比+10%以上なら維持"),
        ("月", "9:30", "エラーログの確認 (Sentry / Vercel)", "エンジニア", "毎週", "Critical 0件を維持"),
        ("月", "10:00", "工務店10社への週次レポート送信", "営業", "毎週", "全社送信完了で✓"),
        ("火", "10:00", "問合せフォームから来た連絡への対応", "営業", "毎日", "24h以内に全件返信"),
        ("水", "10:00", "YouTube 新着動画の公開 + ぺいほーむ掲載", "代表", "毎週水", "YouTube公開から1時間以内に掲載"),
        ("水", "14:00", "SNS: 動画公開の拡散 (X, Instagram, Threads)", "代表", "毎週水", "全プラットフォーム投稿で✓"),
        ("金", "10:00", "見学会予約の工務店確認 (通知届いたか)", "営業", "毎週", "全件確認で✓"),
        ("金", "15:00", "工務店10社のうち1社と電話フォローアップ", "営業", "毎週", "週1社 × 10週 = 全社巡回"),
        ("金", "17:00", "週次 KPI を経理・代表に共有", "PM", "毎週", "月次売上ペースも確認"),
    ]:
        R(ws, row, list(item))
        row += 1
    ws.freeze_panes = "A3"

    # ── ③ Phase別機能追加 ──
    ws = wb.create_sheet("③ Phase別機能追加")
    row = T(ws, 1, "③ Phase別 機能追加スケジュール", 7, "いつ何を復活させるか")
    H(ws, row, ["Phase", "期間", "追加機能", "対応パス", "担当", "期日", "収益モデルへの影響"], [8, 20, 26, 32, 14, 14, 38])
    row += 1
    for f in [
        ("P1", "2026/05/01〜06/30", "運用のみ (MVP 17画面)", "−", "運用チーム", "2026-06-30", "見学会予約 ¥50,000/件 のみ運用開始"),
        ("P2", "2026/07/01〜08/31", "業界ニュース", "/news, /news/[id]", "エンジニア", "2026-07-01", "−"),
        ("P2", "2026/07/01〜08/31", "記事・コラム", "/articles, /articles/[id]", "エンジニア", "2026-07-01", "−"),
        ("P2", "2026/07/01〜08/31", "管理画面 ニュース/記事CMS", "/admin/news, /admin/articles", "エンジニア", "2026-07-03", "−"),
        ("P2", "2026/07/01〜08/31", "B2B ニュース", "/biz/news", "エンジニア", "2026-07-15", "−"),
        ("P2", "2026/07/01〜08/31", "撮影プラン申込フォーム", "内部実装", "エンジニア", "2026-08-15", "撮影費 ¥150,000/本 開始"),
        ("P3", "2026/09/01〜11/30", "販売中の分譲戸建", "/sale-homes, /sale-homes/[id]", "エンジニア", "2026-09-01", "月額掲載 ¥30,000/¥80,000 開始"),
        ("P3", "2026/09/01〜11/30", "取扱中の土地", "/lands, /lands/[id]", "エンジニア", "2026-09-01", "−"),
        ("P3", "2026/09/01〜11/30", "特集", "/features, /features/[id]", "エンジニア", "2026-09-01", "−"),
        ("P3", "2026/09/01〜11/30", "ローンシミュレーター", "/simulator", "エンジニア", "2026-09-05", "−"),
        ("P3", "2026/09/01〜11/30", "成果報酬トラッキング実装", "lead_source tracking", "エンジニア", "2026-09-01", "★成果報酬 3% 開始 (無料相談/AI診断/会員成約)"),
        ("P3", "2026/09/01〜11/30", "エリア別ページ", "/area", "エンジニア", "2026-09-10", "−"),
        ("P3", "2026/09/01〜11/30", "工務店 建売/土地CMS", "/dashboard/builder/sale-homes, /lands", "エンジニア", "2026-09-01", "−"),
        ("P3", "2026/09/01〜11/30", "工務店比較", "/builders/compare", "エンジニア", "2026-10-01", "−"),
        ("P4", "2026/12/01〜2027/02/28", "匿名AI仲介質問", "/mypage/questions", "エンジニア", "2026-12-01", "SaaS ¥50,000/¥150,000 開始"),
        ("P4", "2026/12/01〜2027/02/28", "ウェビナー", "/webinar, /webinar/[id]", "エンジニア", "2026-12-15", "SaaS に含む"),
        ("P4", "2026/12/01〜2027/02/28", "マガジン", "/magazine", "エンジニア", "2027-01-15", "SaaS に含む"),
        ("P4", "2026/12/01〜2027/02/28", "インタビュー", "/interview, /interview/[id]", "エンジニア", "2027-01-15", "SaaS に含む"),
        ("P4", "2026/12/01〜2027/02/28", "Stripe決済 + 請求自動化", "内部実装", "エンジニア", "2026-11-30", "全プランに必須"),
        ("P4", "2026/12/01〜2027/02/28", "管理画面 全CMS", "/admin/* 全て", "エンジニア", "2027-01-31", "−"),
        ("P5", "2027/03〜", "全国工務店追加", "builders-data 拡張", "営業+エンジニア", "2027-03-31", "有料広告投資開始"),
    ]:
        R(ws, row, list(f))
        row += 1
    ws.freeze_panes = "A3"

    # ── ④ 工務店10社 対応管理 ──
    ws = wb.create_sheet("④ 工務店10社 対応管理")
    row = T(ws, 1, "④ 工務店10社 対応管理", 12, "全員同一料金。10社は AI診断の優先送客のみ差別化")
    H(ws, row, [
        "#", "社名", "担当", "契約書", "MVP説明", "撮影予定", "掲載URL",
        "累計予約数", "累計撮影本数", "累計成約数", "最終連絡日", "次回アクション",
    ], [4, 22, 12, 10, 10, 12, 28, 10, 12, 10, 12, 30])
    row += 1
    for i in range(1, 11):
        R(ws, row, [
            i, f"(社名 #{i})", "", "未締結", "未実施", "未定", "",
            0, 0, 0, "", "",
        ])
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑤ リリース前チェックリスト ──
    ws = wb.create_sheet("⑤ リリース前チェックリスト")
    row = T(ws, 1, "⑤ リリース前チェックリスト", 5, "4/30 までに ✓ を全て埋める")
    H(ws, row, ["#", "カテゴリ", "項目", "責任者", "完了 ✓"], [5, 14, 55, 14, 10])
    row += 1
    for r in [
        ("1", "インフラ", "Vercel 本番プロジェクト作成完了", "エンジニア", ""),
        ("2", "インフラ", "独自ドメイン payhome.jp SSL 有効", "エンジニア", ""),
        ("3", "インフラ", "Supabase 本番環境マイグレーション完了", "エンジニア", ""),
        ("4", "インフラ", "本番 .env 全キー設定済み", "エンジニア", ""),
        ("5", "セキュリティ", "NEXTAUTH_SECRET 本番用を新規生成", "エンジニア", ""),
        ("6", "セキュリティ", "Turnstile 有効 (signup/contact/event)", "エンジニア", ""),
        ("7", "セキュリティ", "管理画面 admin ロール認証動作確認", "エンジニア", ""),
        ("8", "セキュリティ", "工務店ダッシュボード builder ロール確認", "エンジニア", ""),
        ("9", "計測", "GA4 + lead_source tracking 動作確認", "エンジニア", ""),
        ("10", "計測", "Sentry エラー監視セットアップ", "エンジニア", ""),
        ("11", "計測", "Vercel Analytics 有効", "エンジニア", ""),
        ("12", "テスト", "17公開ページ 200 確認", "エンジニア", ""),
        ("13", "テスト", "40+非公開ページ 404 確認", "エンジニア", ""),
        ("14", "テスト", "AI診断フロー完走確認 (10社優先表示)", "エンジニア", ""),
        ("15", "テスト", "見学会予約フロー完走確認", "エンジニア", ""),
        ("16", "テスト", "会員登録・ログイン動作確認", "エンジニア", ""),
        ("17", "テスト", "モバイル表示 (iPhone, Android) 確認", "エンジニア", ""),
        ("18", "SEO", "robots.txt 本番URLで閲覧可能", "エンジニア", ""),
        ("19", "SEO", "sitemap.xml 本番URLで閲覧可能", "エンジニア", ""),
        ("20", "SEO", "Search Console 登録 + sitemap 送信", "PM", ""),
        ("21", "SEO", "OGP 画像 + Twitter Card 動作確認", "エンジニア", ""),
        ("22", "営業", "10社と料金プラン書面締結 (¥50,000/件)", "営業", ""),
        ("23", "営業", "10社の掲載情報最終チェック", "営業", ""),
        ("24", "営業", "営業FAQ 20問 全スタッフ共有", "営業", ""),
        ("25", "営業", "リリース予告メール10社送信", "営業", ""),
        ("26", "オペレーション", "問合せ対応フロー決定", "PM", ""),
        ("27", "オペレーション", "見学会予約 → 工務店通知ルート動作確認", "PM", ""),
        ("28", "経理", "請求書テンプレ作成", "経理", ""),
        ("29", "経理", "振込先確認・社内共有", "経理", ""),
        ("30", "広報", "プレスリリース原稿承認", "代表", ""),
        ("31", "広報", "YouTube 告知動画 公開予約", "代表", ""),
        ("32", "広報", "X/Instagram/Threads 告知予約", "代表", ""),
        ("33", "リリース", "リリースリハーサル完了", "エンジニア", ""),
        ("34", "リリース", "ロールバック手順確認", "エンジニア", ""),
    ]:
        R(ws, row, list(r))
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑥ 月次KPI (新料金モデル) ──
    ws = wb.create_sheet("⑥ 月次KPI")
    row = T(ws, 1, "⑥ 月次 KPI トラッキング (全員同一料金)", 14, "Phase 1〜5 の月次KPI目標。実績は毎月末に記入")
    months = [p["month"] for p in PROJECTIONS]
    H(ws, row, ["指標", "Phase"] + months, [18, 6] + [10] * 12)
    row += 1

    # Phase row
    R(ws, row, ["-", "-"] + [p["phase"] for p in PROJECTIONS], fill=CREAM, align=CENTER)
    row += 1

    # 見学会予約件数
    R(ws, row, ["見学会予約 件数 (目標)", ""] + [p["booking_n"] for p in PROJECTIONS], fill=CREAM, align=CENTER)
    row += 1
    R(ws, row, ["  実績", ""] + [""] * 12, fill=GREEN, align=CENTER)
    row += 1

    # 撮影本数
    R(ws, row, ["撮影プラン 本数", ""] + [p["filming_n"] for p in PROJECTIONS], fill=CREAM, align=CENTER)
    row += 1
    R(ws, row, ["  実績", ""] + [""] * 12, fill=GREEN, align=CENTER)
    row += 1

    # 月額掲載 社数
    R(ws, row, ["月額掲載 社数", ""] + [p["sub_n"] for p in PROJECTIONS], fill=CREAM, align=CENTER)
    row += 1
    R(ws, row, ["  実績", ""] + [""] * 12, fill=GREEN, align=CENTER)
    row += 1

    # 成果報酬 件数
    R(ws, row, ["成果報酬 成約件数", ""] + [p["success_n"] for p in PROJECTIONS], fill=CREAM, align=CENTER)
    row += 1
    R(ws, row, ["  実績", ""] + [""] * 12, fill=GREEN, align=CENTER)
    row += 1

    # SaaS 社数
    R(ws, row, ["SaaS 総社数", ""] + [p["saas_lite_n"] + p["saas_pro_n"] for p in PROJECTIONS], fill=CREAM, align=CENTER)
    row += 1
    R(ws, row, ["  実績", ""] + [""] * 12, fill=GREEN, align=CENTER)
    row += 1

    # 月次売上
    R(ws, row, ["月次売上 (円)", ""] + [f"¥{p['total']:,}" for p in PROJECTIONS], fill=CREAM, align=CENTER, bold=True)
    row += 1
    R(ws, row, ["  実績", ""] + [""] * 12, fill=GREEN, align=CENTER)
    row += 1

    ws.freeze_panes = "C3"

    # ── ⑦ 非公開パス復活 ──
    ws = wb.create_sheet("⑦ 非公開パス復活")
    row = T(ws, 1, "⑦ 非公開パス復活スケジュール", 6, "middleware.ts の HIDDEN_PATH_PREFIXES から削除するタイミング")
    H(ws, row, ["#", "パス", "Phase", "復活予定", "実装担当", "備考"], [5, 40, 8, 14, 14, 38])
    row += 1
    for r in [
        (1, "/news", "P2", "2026-07-01", "エンジニア", "業界ニュース一覧・詳細"),
        (2, "/articles", "P2", "2026-07-01", "エンジニア", "記事・コラム"),
        (3, "/admin/news", "P2", "2026-07-03", "エンジニア", "ニュースCMS"),
        (4, "/admin/articles", "P2", "2026-07-03", "エンジニア", "記事CMS"),
        (5, "/biz/news", "P2", "2026-07-15", "エンジニア", "B2B業界ニュース"),
        (6, "/biz/articles", "P2", "2026-08-01", "エンジニア", "B2B集客ノウハウ"),
        (7, "/sale-homes", "P3", "2026-09-01", "エンジニア", "建売一覧・詳細"),
        (8, "/lands", "P3", "2026-09-01", "エンジニア", "土地一覧・詳細"),
        (9, "/features", "P3", "2026-09-01", "エンジニア", "特集"),
        (10, "/simulator", "P3", "2026-09-05", "エンジニア", "ローンシミュレーター"),
        (11, "/area", "P3", "2026-09-10", "エンジニア", "エリア別ページ"),
        (12, "/dashboard/builder/sale-homes", "P3", "2026-09-01", "エンジニア", "工務店 建売CMS"),
        (13, "/dashboard/builder/lands", "P3", "2026-09-01", "エンジニア", "工務店 土地CMS"),
        (14, "/admin/sale-homes", "P3", "2026-09-01", "エンジニア", "管理 建売CMS"),
        (15, "/admin/lands", "P3", "2026-09-01", "エンジニア", "管理 土地CMS"),
        (16, "/admin/features", "P3", "2026-09-01", "エンジニア", "管理 特集CMS"),
        (17, "/builders/compare", "P3", "2026-10-01", "エンジニア", "工務店比較ページ"),
        (18, "/mypage/questions", "P4", "2026-12-01", "エンジニア", "匿名AI仲介質問"),
        (19, "/dashboard/builder/questions", "P4", "2026-12-01", "エンジニア", "工務店Q&A"),
        (20, "/admin/reports", "P4", "2026-12-01", "エンジニア", "月次レポート自動集計"),
        (21, "/webinar", "P4", "2026-12-15", "エンジニア", "ウェビナー一覧・詳細"),
        (22, "/interview", "P4", "2027-01-15", "エンジニア", "工務店インタビュー"),
        (23, "/magazine", "P4", "2027-01-15", "エンジニア", "マガジン"),
        (24, "/biz/webinar", "P4", "2026-12-15", "エンジニア", "B2Bウェビナー"),
        (25, "/dashboard/builder/billing", "P4", "2026-12-01", "エンジニア", "工務店 請求・プラン管理 (Stripe)"),
        (26, "/biz/service", "P4", "2026-12-01", "エンジニア", "B2Bサービス概要"),
        (27, "/biz/ad", "P4", "2026-12-01", "エンジニア", "B2B広告・タイアップ"),
        (28, "/biz/partner", "P4", "2026-12-01", "エンジニア", "B2B提携パートナー募集"),
        (29, "/dashboard/builder/case-studies", "P4", "2027-01-01", "エンジニア", "工務店 施工事例CMS"),
        (30, "/dashboard/builder/videos", "P4", "2027-01-01", "エンジニア", "工務店 動画CMS"),
        (31, "/dashboard/builder/reviews", "P4", "2027-01-01", "エンジニア", "工務店 お客様の声CMS"),
        (32, "/admin/interviews", "P4", "2027-01-15", "エンジニア", "管理 インタビューCMS"),
        (33, "/admin/webinars", "P4", "2026-12-15", "エンジニア", "管理 ウェビナーCMS"),
        (34, "/admin/magazine", "P4", "2027-01-15", "エンジニア", "管理 マガジンCMS"),
        (35, "/admin/videos", "P4", "2027-01-01", "エンジニア", "管理 動画CMS"),
        (36, "/admin/case-studies", "P4", "2027-01-01", "エンジニア", "管理 施工事例CMS"),
        (37, "/admin/reviews", "P4", "2027-01-01", "エンジニア", "管理 お客様の声CMS"),
        (38, "/admin/biz-articles", "P4", "2027-01-15", "エンジニア", "管理 B2B記事CMS"),
        (39, "/admin/biz-news", "P4", "2027-01-15", "エンジニア", "管理 B2BニュースCMS"),
        (40, "/admin/biz-webinars", "P4", "2027-01-15", "エンジニア", "管理 B2BウェビナーCMS"),
        (41, "/admin/users", "P4", "2027-02-01", "エンジニア", "ユーザー管理"),
        (42, "/admin/properties", "P4", "2027-02-01", "エンジニア", "物件データ管理"),
        (43, "/admin/security/activity/system/notifications/data", "P4", "2027-02-15", "エンジニア", "システム運用系画面"),
        (44, "/mypage/catalog", "P4", "2027-01-01", "エンジニア", "マイページのカタログダウンロード"),
        (45, "/mypage/feedback", "P4", "2027-02-01", "エンジニア", "マイページ フィードバック"),
        (46, "/welcome, /property, /builders/contact", "P5", "2027-03-01", "エンジニア", "全国展開時に再検討"),
    ]:
        R(ws, row, list(r))
        row += 1
    ws.freeze_panes = "A3"

    out = OUT_MGMT / "MVP実行計画.xlsx"
    wb.save(out)
    return out


# ══════════════════════════════════════════════════════════════════
# ② フェーズ別収益モデル.xlsx を再生成
# ══════════════════════════════════════════════════════════════════

def build_revenue_model() -> Path:
    wb = Workbook()

    # 目次
    ws = wb.active
    ws.title = "目次"
    row = T(ws, 1, "ぺいほーむ フェーズ別 収益モデル (全員同一料金)", 3, "10社料金優遇ゼロ。差別化はAI診断での優先送客のみ")
    H(ws, row, ["#", "シート名", "用途"], [6, 30, 60])
    row += 1
    for t in [
        ("①", "料金プラン一覧", "Phase 1〜5 の料金プラン (全員同一)"),
        ("②", "機能x料金マトリクス", "どの機能がどの料金プランで使えるか"),
        ("③", "月次売上シミュレーション", "12ヶ月の想定売上と内訳"),
        ("④", "工務店別プラン管理", "10社 + 新規それぞれの適用プラン (料金は全員同一)"),
        ("⑤", "Phase移行 判断基準", "次のPhaseへ進む条件"),
        ("⑥", "プラン別の原価", "プランごとの粗利率"),
    ]:
        R(ws, row, list(t))
        row += 1

    # ── ① 料金プラン一覧 ──
    ws = wb.create_sheet("① 料金プラン一覧")
    row = T(ws, 1, "① Phase別 料金プラン一覧 (全員同一)", 5, "10社と新規で料金は同じ。10社優遇はAI診断優先送客のみ")
    H(ws, row, ["Phase", "項目", "料金 (税別)", "課金タイミング", "備考"], [8, 28, 18, 20, 40])
    row += 1
    for p in [
        ("P1", "見学会予約", "¥50,000/件", "予約成立時", "5/1から課金開始、全員同一"),
        ("P1", "成果報酬", "−", "−", "Phase 3 から導入"),
        ("P1", "AI家づくり診断", "無料", "−", "ユーザー無料"),
        ("P2", "見学会予約 (継続)", "¥50,000/件", "予約成立時", "−"),
        ("P2", "撮影プラン", "¥150,000/本", "撮影日", "ルームツアー撮影+編集+YouTube掲載"),
        ("P3", "見学会予約 (継続)", "¥50,000/件", "予約成立時", "−"),
        ("P3", "撮影プラン (継続)", "¥150,000/本", "撮影日", "−"),
        ("P3", "月額掲載 ベーシック", "¥30,000/月", "毎月月末請求", "物件3件まで"),
        ("P3", "月額掲載 プロ", "¥80,000/月", "毎月月末請求", "物件無制限+月1本特集枠"),
        ("P3", "成果報酬 3%", "契約額の 3% (上限なし)", "契約成立時", "★無料相談/AI診断/会員登録済みユーザーの成約のみ"),
        ("P4", "見学会予約 (継続)", "¥50,000/件", "予約成立時", "−"),
        ("P4", "撮影プラン (継続)", "¥150,000/本", "撮影日", "−"),
        ("P4", "月額掲載 (継続)", "¥30,000/月 / ¥80,000/月", "毎月月末請求", "−"),
        ("P4", "成果報酬 (継続)", "3%", "契約成立時", "−"),
        ("P4", "SaaS ライト", "¥50,000/月", "毎月月末請求", "AI質問無制限+インタビュー掲載+月次レポート"),
        ("P4", "SaaS プロ", "¥150,000/月", "毎月月末請求", "全機能+専任担当+API連携"),
        ("P5", "全項目継続", "正規価格", "各タイミング", "全国展開+有料広告投資"),
    ]:
        R(ws, row, list(p))
        row += 1
    ws.freeze_panes = "A3"

    # ── ② 機能x料金マトリクス ──
    ws = wb.create_sheet("② 機能x料金")
    row = T(ws, 1, "② 機能 x 料金プラン マトリクス", 7, "どの機能がどの料金プランで提供されるか (全員同一)")
    H(ws, row, ["機能", "無料", "見学会予約", "撮影プラン", "月額ベーシック", "月額プロ", "SaaS"], [30, 10, 12, 12, 16, 12, 10])
    row += 1
    for m in [
        ("ぺいほーむ掲載 (基本情報)", "✓", "✓", "✓", "✓", "✓", "✓"),
        ("AI家づくり診断 送客", "✓", "✓", "✓", "✓", "✓", "✓"),
        ("見学会予約 受付", "−", "✓", "✓", "✓", "✓", "✓"),
        ("ルームツアー動画 掲載", "−", "−", "✓ (撮影時)", "−", "−", "−"),
        ("物件 3件掲載", "−", "−", "−", "✓", "✓", "✓"),
        ("物件 無制限掲載", "−", "−", "−", "−", "✓", "✓"),
        ("特集枠 月1本", "−", "−", "−", "−", "✓", "✓"),
        ("AI質問 対応 (無制限)", "−", "−", "−", "−", "−", "✓"),
        ("インタビュー 掲載", "−", "−", "−", "−", "−", "✓"),
        ("月次レポート 自動配信", "−", "−", "−", "−", "−", "✓"),
        ("API 連携", "−", "−", "−", "−", "−", "Pro のみ"),
        ("専任担当サポート", "−", "−", "−", "−", "−", "Pro のみ"),
    ]:
        R(ws, row, list(m))
        row += 1
    ws.freeze_panes = "A3"

    # ── ③ 月次売上シミュレーション ──
    ws = wb.create_sheet("③ 月次売上シミュレーション")
    row = T(ws, 1, "③ 月次売上シミュレーション (12ヶ月)", 8, "全員同一料金ベース。Phase 3 から成果報酬 (無料相談/AI診断/会員成約対象)")
    H(ws, row, ["月", "Phase", "見学会予約", "撮影", "月額掲載", "成果報酬", "SaaS", "月次合計"], [8, 6, 14, 14, 14, 14, 14, 16])
    row += 1
    for p in PROJECTIONS:
        fill = GRAY if p["total"] < 800_000 else (CREAM if p["total"] < 3_000_000 else GREEN)
        R(ws, row, [
            p["month"], p["phase"],
            f"¥{p['booking_rev']:,}",
            f"¥{p['filming_rev']:,}",
            f"¥{p['sub_rev']:,}",
            f"¥{p['success_rev']:,}",
            f"¥{p['saas_rev']:,}",
            f"¥{p['total']:,}",
        ], fill=fill, align=CENTER)
        row += 1
    total_year = sum(p["total"] for p in PROJECTIONS)
    R(ws, row, ["年間合計", "", "", "", "", "", "", f"¥{total_year:,}"], fill=ORANGE, align=CENTER)
    for c in range(1, 9):
        ws.cell(row=row, column=c).font = WHITE_BOLD
    ws.freeze_panes = "A3"

    # ── ④ 工務店別プラン管理 ──
    ws = wb.create_sheet("④ 工務店別プラン管理")
    row = T(ws, 1, "④ 工務店別プラン管理 (全員同一料金)", 8, "社ごとの適用プランを管理 (料金は全員同じ)")
    H(ws, row, ["#", "社名", "契約日", "分類", "見学会予約", "撮影", "月額掲載", "SaaS"], [4, 22, 12, 14, 14, 14, 18, 14])
    row += 1
    for i in range(1, 11):
        R(ws, row, [
            i, f"(既存10社 #{i})", "2026-04-XX", "既存 (優先送客対象)",
            "¥50,000/件", "¥150,000/本 (P2〜)", "¥30,000/月 (P3〜)", "¥50,000/月 (P4〜)",
        ], fill=CREAM)
        row += 1
    for i in range(11, 16):
        R(ws, row, [
            i, "(新規 ここから)", "", "新規",
            "¥50,000/件", "¥150,000/本 (P2〜)", "¥30,000/月 (P3〜)", "¥50,000/月 (P4〜)",
        ])
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑤ Phase移行 判断基準 ──
    ws = wb.create_sheet("⑤ Phase移行 判断基準")
    row = T(ws, 1, "⑤ Phase 移行 判断基準", 4, "次のフェーズに進むかどうかを決めるKPI")
    H(ws, row, ["Phase", "判断タイミング", "必須条件", "補足条件"], [8, 18, 55, 50])
    row += 1
    for c in [
        ("P0→P1", "2026-04-30", "17公開ページ200・40+非公開404・10社と書面締結 (¥50,000/件)", "プレスリリース準備完了"),
        ("P1→P2", "2026-06-30", "見学会予約が2ヶ月連続で月10件以上", "10社のうち8社が撮影プランに興味あり"),
        ("P2→P3", "2026-08-31", "撮影プラン申込が2ヶ月で3本以上、見学会予約が月20件以上", "記事/ニュースCMSが安定稼働、成果報酬トラッキング実装済み"),
        ("P3→P4", "2026-11-30", "月額掲載の契約5社以上、成果報酬が月1件以上発生、総売上月¥3,000,000 超", "物件情報の登録が安定"),
        ("P4→P5", "2027-02-28", "SaaSプラン契約5社以上、総売上月¥7,000,000 超", "全国展開の準備完了 (マーケ・採用計画)"),
    ]:
        R(ws, row, list(c), align=LEFT_TOP)
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑥ プラン別原価 ──
    ws = wb.create_sheet("⑥ プラン別原価")
    row = T(ws, 1, "⑥ プラン別 原価と粗利 (全員同一料金)", 5, "プラットフォームの採算性評価")
    H(ws, row, ["プラン", "正規価格", "変動原価", "粗利", "粗利率"], [30, 18, 18, 18, 12])
    row += 1
    for name, price, var, profit, rate in [
        ("見学会予約 (件あたり)", 50_000, 5_000, 45_000, "90%"),
        ("撮影プラン (本あたり)", 150_000, 80_000, 70_000, "47%"),
        ("月額掲載 ベーシック", 30_000, 3_000, 27_000, "90%"),
        ("月額掲載 プロ", 80_000, 10_000, 70_000, "88%"),
        ("成果報酬 (¥3000万契約想定)", 900_000, 20_000, 880_000, "98%"),
        ("SaaS ライト", 50_000, 8_000, 42_000, "84%"),
        ("SaaS プロ", 150_000, 30_000, 120_000, "80%"),
    ]:
        R(ws, row, [name, f"¥{price:,}", f"¥{var:,}", f"¥{profit:,}", rate], align=CENTER)
        row += 1
    ws.freeze_panes = "A3"

    out = OUT_MGMT / "フェーズ別収益モデル.xlsx"
    wb.save(out)
    return out


# ══════════════════════════════════════════════════════════════════
# ③ 動線設計シート.xlsx を再生成 (簡略版)
# ══════════════════════════════════════════════════════════════════

def build_funnel_design() -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "目次"
    row = T(ws, 1, "ぺいほーむ 動線設計シート (全員同一料金)", 3, "目標収益から逆算した動線設計とKPI管理")
    H(ws, row, ["#", "シート名", "用途"], [6, 32, 60])
    row += 1
    for t in [
        ("①", "収益→件数 逆算", "Phase別の月次売上目標と必要件数"),
        ("②", "動線マップ 全体像", "ユーザーの入口から収益発生までの動線"),
        ("③", "動線別 Phase1 KPI", "MVP 17画面の動線"),
        ("④", "動線別 Phase2 KPI", "P1 + 記事・ニュース・撮影プラン"),
        ("⑤", "動線別 Phase3 KPI", "P2 + 物件・月額掲載・成果報酬"),
        ("⑥", "動線別 Phase4 KPI", "P3 + AI質問・SaaS"),
        ("⑦", "動線別 Phase5 KPI", "全国展開期"),
        ("⑧", "ファネル逆算", "UU→診断→予約→成約 の各転換率"),
        ("⑨", "動線施策 優先度", "施策優先度マトリクス"),
    ]:
        R(ws, row, list(t))
        row += 1

    # ── ① 収益→件数 逆算 ──
    ws = wb.create_sheet("① 収益→件数 逆算")
    row = T(ws, 1, "① 収益目標から件数への逆算", 8, "Phase別の月次売上目標と必要件数 (全員同一料金)")
    H(ws, row, ["Phase", "月", "売上目標", "見学会予約", "成果報酬件数", "撮影本数", "月額掲載社数", "SaaS社数"], [6, 8, 16, 14, 14, 12, 14, 12])
    row += 1
    for p in PROJECTIONS:
        fill = GRAY if p["total"] < 800_000 else (CREAM if p["total"] < 3_000_000 else GREEN)
        R(ws, row, [
            p["phase"], p["month"], f"¥{p['total']:,}",
            p["booking_n"], p["success_n"], p["filming_n"], p["sub_n"],
            p["saas_lite_n"] + p["saas_pro_n"],
        ], align=CENTER, fill=fill)
        row += 1
    R(ws, row, [
        "年間", "", f"¥{sum(p['total'] for p in PROJECTIONS):,}",
        sum(p["booking_n"] for p in PROJECTIONS),
        sum(p["success_n"] for p in PROJECTIONS),
        sum(p["filming_n"] for p in PROJECTIONS),
        sum(p["sub_n"] for p in PROJECTIONS),
        sum(p["saas_lite_n"] + p["saas_pro_n"] for p in PROJECTIONS),
    ], align=CENTER, fill=ORANGE, bold=True)
    for c in range(1, 9):
        ws.cell(row=row, column=c).font = WHITE_BOLD
    ws.freeze_panes = "A3"

    # ── ② 動線マップ ──
    ws = wb.create_sheet("② 動線マップ")
    row = T(ws, 1, "② 動線マップ (入口→収益)", 5, "ユーザーがどこから入ってどこで収益が発生するか")
    H(ws, row, ["入口", "第1ステップ", "第2ステップ", "ゴール", "収益発生"], [20, 26, 26, 22, 24])
    row += 1
    for r in [
        ("YouTube チャンネル", "動画説明欄リンク", "ぺいほーむTOP or 工務店ページ", "見学会予約", "見学会予約費 ¥50,000"),
        ("Google検索 (指名)", "ぺいほーむTOP", "工務店一覧 → 詳細", "見学会予約", "¥50,000"),
        ("Google検索 (平屋 鹿児島)", "工務店一覧 or 事例", "工務店詳細 → 見学会", "見学会予約", "¥50,000"),
        ("SNS (X/Insta/Threads)", "ぺいほーむTOP", "AI診断", "診断結果 → 予約/問合せ", "¥50,000 + 成果報酬 3% (P3〜)"),
        ("AI家づくり診断", "診断フォーム", "結果ページ", "お気に入り or 予約/相談", "¥50,000 + 成果報酬 3% (P3〜)"),
        ("無料相談", "相談フォーム", "スタッフ対応", "見学会予約 or 契約", "¥50,000 + 成果報酬 3% (P3〜)"),
        ("会員登録", "signup", "マイページ活用 → 工務店連絡", "契約", "成果報酬 3% (P3〜)"),
        ("デジタルカタログDL", "カタログページ", "会員登録", "メルマガ → 診断・予約", "長期的送客 + 成果報酬 (P3〜)"),
        ("Phase2: 業界ニュース", "ニュース記事", "工務店詳細", "見学会予約", "¥50,000"),
        ("Phase3: 建売物件", "物件詳細", "工務店詳細", "物件問合せ", "月額掲載料 + 成果報酬 (会員成約)"),
        ("Phase4: 匿名AI質問", "質問投稿", "工務店回答", "リード → 予約", "SaaSプラン + 成果報酬"),
    ]:
        R(ws, row, list(r))
        row += 1
    ws.freeze_panes = "A3"

    # ── ③-⑦ Phase別動線KPI (共通関数) ──
    def build_phase_sheet(name: str, title: str, subtitle: str, routes: list):
        ws = wb.create_sheet(name)
        row = T(ws, 1, title, 9, subtitle)
        H(ws, row, [
            "動線名", "入口", "主要KPI", "月間目標(UU)",
            "転換率", "月次リード", "見学会予約", "収益インパクト", "担当",
        ], [20, 22, 22, 14, 12, 14, 14, 20, 12])
        row += 1
        for r in routes:
            R(ws, row, list(r), align=LEFT_TOP)
            row += 1
        ws.freeze_panes = "A3"

    p1 = [
        ("YouTube動線", "YouTubeチャンネル", "動画視聴数→遷移率", "3,000 UU", "5%", "150", "10件", "¥500,000/月", "代表"),
        ("AI診断動線", "TOPのAI診断CTA", "診断完了数", "200 UU", "30%", "60", "3件", "¥150,000/月", "PM"),
        ("工務店検索動線", "/builders", "工務店詳細閲覧", "500 UU", "8%", "40", "2件", "¥100,000/月", "営業"),
        ("見学会動線", "/event", "予約フォーム完了率", "200 UU", "15%", "30", "3件", "¥150,000/月", "営業"),
        ("無料相談動線", "TOP相談CTA", "相談フォーム送信", "100 UU", "5%", "5", "1件", "¥50,000/月 (P3〜は成果報酬対象)", "営業"),
        ("事例ライブラリ動線", "/case-studies", "事例→工務店遷移率", "300 UU", "10%", "30", "1件", "¥50,000/月", "PM"),
    ]
    build_phase_sheet("③ Phase1 動線KPI", "③ Phase 1 動線別KPI (2026/05/01〜06/30)", "見学会予約 ¥50,000/件 のみ発生", p1)

    p2 = p1 + [
        ("業界ニュース動線", "/news", "ニュース→工務店遷移", "800 UU", "10%", "80", "2件", "¥100,000/月", "PM"),
        ("記事動線", "/articles", "記事→工務店遷移", "1,200 UU", "8%", "96", "3件", "¥150,000/月", "PM"),
        ("撮影プラン動線", "工務店アウトバウンド", "撮影プラン申込", "10社", "20%", "2社", "−", "¥300,000/月", "代表"),
    ]
    build_phase_sheet("④ Phase2 動線KPI", "④ Phase 2 動線別KPI (2026/07/01〜08/31)", "撮影プラン開始", p2)

    p3 = p2 + [
        ("建売物件動線", "/sale-homes", "物件→問合せ", "1,500 UU", "6%", "90", "4件", "¥200,000/月", "PM"),
        ("土地情報動線", "/lands", "土地→問合せ", "800 UU", "5%", "40", "2件", "¥100,000/月", "PM"),
        ("特集動線", "/features", "特集→工務店", "1,000 UU", "10%", "100", "3件", "¥150,000/月", "PM"),
        ("月額掲載動線", "工務店アウトバウンド", "月額プラン加入", "15社", "30%", "5社", "−", "¥150,000/月", "営業"),
        ("成果報酬動線", "無料相談/AI診断経由", "成約数", "60リード", "2%", "1成約", "−", "¥900,000/件", "PM+営業"),
    ]
    build_phase_sheet("⑤ Phase3 動線KPI", "⑤ Phase 3 動線別KPI (2026/09/01〜11/30)", "月額掲載 + 成果報酬 開始", p3)

    p4 = p3 + [
        ("AI質問動線", "/mypage/questions", "質問投稿→回答率", "200 質問", "80%", "160", "4件", "¥200,000/月", "PM"),
        ("ウェビナー動線", "/webinar", "参加申込→成約", "500 UU", "8%", "40", "3件", "¥150,000/月", "PM"),
        ("SaaS動線", "工務店アウトバウンド", "SaaS加入", "20社", "25%", "5社", "−", "¥500,000/月", "営業"),
    ]
    build_phase_sheet("⑥ Phase4 動線KPI", "⑥ Phase 4 動線別KPI (2026/12/01〜2027/02/28)", "SaaS開始、MRR成長期", p4)

    p5 = [
        ("県別SEO動線 (福岡)", "/area/fukuoka", "検索流入→工務店", "3,000 UU", "10%", "300", "15件", "¥750,000/月", "PM"),
        ("県別SEO動線 (他県)", "/area/*", "検索流入→工務店", "5,000 UU", "8%", "400", "20件", "¥1,000,000/月", "PM"),
        ("既存動線 (合算)", "P1〜P4", "継続運用", "40,000 UU", "−", "−", "50件", "¥3,000,000/月", "全員"),
        ("有料広告動線", "Google/Meta Ads", "広告CPA", "5,000 UU", "5%", "250", "10件", "¥500,000/月", "代表"),
    ]
    build_phase_sheet("⑦ Phase5 動線KPI", "⑦ Phase 5 動線別KPI (2027/03〜)", "全国展開", p5)

    # ── ⑧ ファネル逆算 ──
    ws = wb.create_sheet("⑧ ファネル逆算")
    row = T(ws, 1, "⑧ ファネル逆算 (詳細)", 6, "UU→診断→会員登録→見学会予約→成約 の各転換率")
    H(ws, row, ["Phase", "目標予約数/月", "必要成約件数", "必要会員登録数", "必要AI診断数", "必要UU"], [6, 18, 16, 16, 16, 16])
    row += 1
    for phase, b, s, req_member, req_diag, req_uu in [
        ("P1", 10, 0, 100, 250, 6_250),
        ("P2", 25, 0, 200, 500, 12_500),
        ("P3", 40, 2, 400, 1_000, 25_000),
        ("P4", 55, 4, 600, 1_375, 34_375),
        ("P5", 70, 5, 900, 1_750, 43_750),
    ]:
        R(ws, row, [phase, f"{b}件", f"{s}件", f"{req_member:,}", f"{req_diag:,}", f"{req_uu:,}"], align=CENTER, fill=CREAM)
        row += 1
    row += 1
    c = ws.cell(row=row, column=1, value="前提転換率")
    c.font = BOLD
    row += 1
    H(ws, row, ["ステップ", "転換率", "備考"], [32, 14, 50])
    row += 1
    for s in [
        ("UU → AI診断", "4%", "TOP CTA からの誘導"),
        ("AI診断 → 工務店詳細", "50%", "診断結果の工務店カードから"),
        ("工務店詳細 → 見学会予約", "10%", "見学会予約CTAの完了率"),
        ("UU → 会員登録", "1.5%", "※ P3 から成果報酬の対象判定に重要"),
        ("会員登録 → 成約", "0.5%", "長期的"),
    ]:
        R(ws, row, list(s))
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑨ 施策優先度 ──
    ws = wb.create_sheet("⑨ 施策優先度")
    row = T(ws, 1, "⑨ 動線施策 優先度マトリクス", 5, "収益インパクト x 実装難易度")
    H(ws, row, ["優先度", "施策", "動線", "収益インパクト", "実装難易度"], [10, 40, 20, 18, 14])
    row += 1
    fill_map = {"S": RED, "A": YELLOW, "B": CREAM, "C": GRAY, "D": BLUE}
    for priority, action, route, impact, diff in [
        ("S", "YouTube動画の冒頭でぺいほーむ紹介を毎本挿入", "YouTube動線", "高", "低"),
        ("S", "見学会予約CTAをモバイルで追従フッター固定", "見学会動線", "高", "低"),
        ("S", "AI診断完了後の工務店カードに見学会予約ボタン併設", "AI診断動線", "高", "低"),
        ("S", "会員登録をAI診断完了時に強めに促す (Phase 3成果報酬対象)", "AI診断→会員", "高", "低"),
        ("A", "動画カードにCTAオーバーレイで『この家が気になる』", "YouTube動線", "高", "中"),
        ("A", "TOP の1st view に見学会予約CTA併設", "見学会動線", "高", "中"),
        ("A", "AI診断の質問数を10→7問に減らす (離脱率改善)", "AI診断動線", "中", "中"),
        ("A", "工務店詳細ページに見学会カレンダー埋込", "工務店検索動線", "高", "中"),
        ("A", "lead_source tracking 実装 (成果報酬の追跡基盤)", "全動線", "高", "中"),
        ("B", "無料相談フォームを1画面でLINE連携", "無料相談動線", "中", "中"),
        ("B", "事例カードに『同じ工務店の動画を見る』", "事例動線", "中", "低"),
        ("B", "Phase2: ニュース記事内のCTA実装", "記事動線", "中", "中"),
        ("B", "Phase2: 撮影プラン申込フォーム導線整備", "撮影動線", "高", "中"),
        ("C", "Phase3: 物件詳細『近くの工務店』レコメンド", "物件動線", "中", "高"),
        ("C", "Phase3: 月額プランの体験版 (1ヶ月無料)", "月額動線", "中", "低"),
        ("C", "Phase3: 成約報告フォーム (工務店ダッシュボード)", "成果報酬動線", "高", "中"),
        ("C", "Phase4: AI質問のリアルタイム通知Push", "AI質問動線", "中", "高"),
        ("C", "Phase4: SaaSデモ環境のセルフサインアップ", "SaaS動線", "高", "高"),
        ("D", "Phase5: 県別ランディングページ x 5県", "県別SEO", "高", "高"),
        ("D", "Phase5: Google広告のリマーケティング", "有料広告", "中", "中"),
    ]:
        R(ws, row, [priority, action, route, impact, diff], fill=fill_map.get(priority))
        row += 1
    ws.freeze_panes = "A3"

    out = OUT_MGMT / "動線設計シート.xlsx"
    wb.save(out)
    return out


# ══════════════════════════════════════════════════════════════════
# ④ 請求・P&L管理.xlsx を再生成 (新料金モデル)
# ══════════════════════════════════════════════════════════════════

def build_accounting() -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "目次"
    row = T(ws, 1, "ぺいほーむ 請求・P&L 管理 (全員同一料金)", 3, "全員同一料金モデル、成果報酬 Phase 3〜")
    H(ws, row, ["#", "シート名", "用途"], [6, 32, 55])
    row += 1
    for t in [
        ("①", "月次請求台帳", "毎月の請求明細"),
        ("②", "入金管理", "請求に対する入金確認"),
        ("③", "月次P&L 12ヶ月", "Phase 1〜5 の月次P&L想定"),
        ("④", "原価・変動費", "プラン別の原価と粗利率"),
        ("⑤", "固定費 年間計画", "人件費・SaaS・家賃等"),
        ("⑥", "税務・請求フロー", "請求書発行から入金確認までの手順"),
    ]:
        R(ws, row, list(t))
        row += 1

    # ── ① 月次請求台帳 ──
    ws = wb.create_sheet("① 月次請求台帳")
    row = T(ws, 1, "① 月次請求台帳", 9, "毎月月末に当月分の請求を記載")
    H(ws, row, ["請求月", "請求日", "請求先", "項目", "単価", "数量", "小計", "税別/込", "ステータス"], [10, 12, 22, 20, 12, 8, 14, 10, 12])
    row += 1
    for s in [
        ("2026-05", "", "(見学会予約発生の工務店)", "見学会予約 ¥50,000/件", 50000, 0, 0, "税別", "未請求"),
        ("2026-06", "", "", "見学会予約 ¥50,000/件", 50000, 0, 0, "税別", "未請求"),
        ("2026-07", "", "", "撮影プラン ¥150,000/本", 150000, 0, 0, "税別", "未請求"),
        ("2026-09", "", "", "月額掲載 ベーシック ¥30,000/月", 30000, 0, 0, "税別", "未請求"),
        ("2026-09", "", "", "月額掲載 プロ ¥80,000/月", 80000, 0, 0, "税別", "未請求"),
        ("2026-09", "", "", "成果報酬 3% (無料相談/AI診断/会員成約)", 0, 0, 0, "税別", "未請求"),
        ("2026-12", "", "", "SaaS ライト ¥50,000/月", 50000, 0, 0, "税別", "未請求"),
        ("2026-12", "", "", "SaaS プロ ¥150,000/月", 150000, 0, 0, "税別", "未請求"),
    ]:
        R(ws, row, list(s), fill=GRAY if s[-1] != "未請求" else None)
        row += 1
    for _ in range(20):
        R(ws, row, [""] * 9)
        row += 1
    ws.freeze_panes = "A3"

    # ── ② 入金管理 ──
    ws = wb.create_sheet("② 入金管理")
    row = T(ws, 1, "② 入金管理", 8, "請求に対する入金の確認")
    H(ws, row, ["請求月", "請求先", "項目", "請求額", "支払期限", "入金日", "入金額", "差額"], [10, 22, 20, 14, 12, 12, 14, 12])
    row += 1
    for _ in range(30):
        R(ws, row, [""] * 8)
        row += 1
    ws.freeze_panes = "A3"

    # ── ③ 月次P&L (12ヶ月計画) ──
    ws = wb.create_sheet("③ 月次P&L")
    row = T(ws, 1, "③ 月次 P&L (12ヶ月計画、全員同一料金)", 14, "毎月実績を記入")
    months = [p["month"] for p in PROJECTIONS]
    H(ws, row, ["項目", "Phase"] + months, [24, 6] + [11] * 12)
    row += 1

    # Phase row
    R(ws, row, ["-", "-"] + [p["phase"] for p in PROJECTIONS], fill=CREAM, align=CENTER)
    row += 1

    # 売上
    for label, key in [
        ("【売上】見学会予約", "booking_rev"),
        ("【売上】撮影プラン", "filming_rev"),
        ("【売上】月額掲載", "sub_rev"),
        ("【売上】成果報酬", "success_rev"),
        ("【売上】SaaS", "saas_rev"),
    ]:
        R(ws, row, [label, ""] + [f"¥{p[key]:,}" for p in PROJECTIONS], fill=CREAM, align=RIGHT)
        ws.cell(row=row, column=1).alignment = LEFT
        row += 1

    R(ws, row, ["売上合計", ""] + [f"¥{p['total']:,}" for p in PROJECTIONS], fill=ORANGE, align=RIGHT, bold=True)
    for c in range(1, 15):
        ws.cell(row=row, column=c).font = WHITE_BOLD
    ws.cell(row=row, column=1).alignment = LEFT
    row += 2

    # 変動原価
    cost_rates = {
        "booking_rev": 0.10,
        "filming_rev": 0.53,
        "sub_rev": 0.10,
        "success_rev": 0.02,
        "saas_rev": 0.16,
    }
    cost_labels = {
        "booking_rev": "【原価】見学会予約 (10%)",
        "filming_rev": "【原価】撮影プラン (53%)",
        "sub_rev": "【原価】月額掲載 (10%)",
        "success_rev": "【原価】成果報酬 (2%)",
        "saas_rev": "【原価】SaaS (16%)",
    }
    cost_totals = []
    for key, rate in cost_rates.items():
        costs = [int(p[key] * rate) for p in PROJECTIONS]
        R(ws, row, [cost_labels[key], ""] + [f"¥{c:,}" for c in costs], align=RIGHT)
        ws.cell(row=row, column=1).alignment = LEFT
        row += 1
        cost_totals.append(costs)

    total_cost = [sum(ct[i] for ct in cost_totals) for i in range(12)]
    R(ws, row, ["変動原価 合計", ""] + [f"¥{c:,}" for c in total_cost], fill=YELLOW, align=RIGHT, bold=True)
    ws.cell(row=row, column=1).alignment = LEFT
    row += 2

    # 粗利
    gross = [PROJECTIONS[i]["total"] - total_cost[i] for i in range(12)]
    R(ws, row, ["粗利", ""] + [f"¥{g:,}" for g in gross], fill=GREEN, align=RIGHT, bold=True)
    ws.cell(row=row, column=1).alignment = LEFT
    row += 1
    R(ws, row, ["粗利率", ""] + [f"{(g/t*100):.0f}%" if t else "−" for g, t in zip(gross, [p["total"] for p in PROJECTIONS])], fill=GREEN, align=RIGHT)
    ws.cell(row=row, column=1).alignment = LEFT
    row += 2

    # 固定費
    fixed_monthly = 600_000
    R(ws, row, ["固定費 (人件費・家賃・SaaS)", ""] + [f"¥{fixed_monthly:,}"] * 12, fill=RED, align=RIGHT)
    ws.cell(row=row, column=1).alignment = LEFT
    row += 1

    # 営業利益
    op = [gross[i] - fixed_monthly for i in range(12)]
    R(ws, row, ["営業利益", ""] + [f"¥{o:,}" for o in op], fill=BLUE, align=RIGHT, bold=True)
    ws.cell(row=row, column=1).alignment = LEFT
    row += 2

    ws.freeze_panes = "C3"

    # ── ④ 原価・変動費 ──
    ws = wb.create_sheet("④ 原価・変動費")
    row = T(ws, 1, "④ プラン別 原価・粗利 (全員同一)", 5, "プラットフォームの採算性評価")
    H(ws, row, ["プラン", "正規価格", "変動原価", "粗利", "粗利率"], [32, 18, 18, 18, 12])
    row += 1
    for name, price, var, profit, rate in [
        ("見学会予約 (1件)", 50_000, 5_000, 45_000, "90%"),
        ("撮影プラン (1本)", 150_000, 80_000, 70_000, "47%"),
        ("月額ベーシック (1社/月)", 30_000, 3_000, 27_000, "90%"),
        ("月額プロ (1社/月)", 80_000, 10_000, 70_000, "88%"),
        ("成果報酬 (¥3000万契約想定)", 900_000, 20_000, 880_000, "98%"),
        ("SaaS ライト (1社/月)", 50_000, 8_000, 42_000, "84%"),
        ("SaaS プロ (1社/月)", 150_000, 30_000, 120_000, "80%"),
    ]:
        R(ws, row, [name, f"¥{price:,}", f"¥{var:,}", f"¥{profit:,}", rate], align=CENTER)
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑤ 固定費 ──
    ws = wb.create_sheet("⑤ 固定費")
    row = T(ws, 1, "⑤ 固定費 年間計画", 4, "毎月発生する固定費の想定")
    H(ws, row, ["項目", "月額", "年額", "備考"], [30, 16, 18, 40])
    row += 1
    fixed = [
        ("代表・PM 人件費", 300_000, 3_600_000, "代表 + PM 1名"),
        ("エンジニア 人件費", 150_000, 1_800_000, "Phase 2から業務委託1名想定"),
        ("営業 人件費", 0, 0, "Phase 1 は代表が兼任"),
        ("Vercel Pro", 3_000, 36_000, "Phase 1 は Hobby"),
        ("Supabase Pro", 4_000, 48_000, "Phase 2 以降"),
        ("Sentry", 3_000, 36_000, ""),
        ("ドメイン", 1_500, 18_000, "payhome.jp"),
        ("経理ソフト (freee)", 3_000, 36_000, ""),
        ("オフィス家賃", 80_000, 960_000, "鹿児島拠点"),
        ("通信費", 10_000, 120_000, ""),
        ("水道光熱費", 15_000, 180_000, ""),
        ("交通費", 20_000, 240_000, "工務店訪問"),
        ("広告宣伝費", 10_000, 120_000, "Phase 1 は YouTube オーガニック"),
    ]
    sum_m = 0
    for name, m, y, note in fixed:
        R(ws, row, [name, f"¥{m:,}", f"¥{y:,}", note])
        sum_m += m
        row += 1
    R(ws, row, ["合計", f"¥{sum_m:,}", f"¥{sum_m*12:,}", ""], fill=ORANGE, bold=True)
    for c in range(1, 4):
        ws.cell(row=row, column=c).font = WHITE_BOLD
    ws.freeze_panes = "A3"

    # ── ⑥ 請求フロー ──
    ws = wb.create_sheet("⑥ 請求フロー")
    row = T(ws, 1, "⑥ 請求書発行から入金確認までの手順", 4, "毎月1日〜月末までのサイクル")
    H(ws, row, ["ステップ", "期日", "担当", "作業内容"], [10, 14, 12, 60])
    row += 1
    for f in [
        ("1", "前月末", "PM", "月次請求台帳 (①シート) に当月発生する請求を記入"),
        ("2", "月初3営業日以内", "経理", "請求書テンプレートで各社の請求書を作成"),
        ("3", "月初3営業日以内", "経理", "請求書を PDF で工務店様へメール送信"),
        ("4", "送信当日", "経理", "入金管理シート (②) に請求行を追加"),
        ("5", "月末", "経理", "入金確認 → 入金管理シートを更新"),
        ("6", "月末翌日", "経理", "未入金案件を PM へエスカレーション"),
        ("7", "月末", "PM", "月次P&L シート (③) を実績値で更新"),
        ("8", "月末翌日", "PM", "代表・経理へ月次レポート共有"),
        ("9", "四半期ごと", "代表・経理", "消費税・法人税の納付準備"),
        ("10", "年度末", "代表・経理", "決算書作成 (外部税理士に依頼)"),
    ]:
        R(ws, row, list(f), align=LEFT_TOP)
        row += 1
    ws.freeze_panes = "A3"

    out = OUT_KEIRI / "請求・P&L管理.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    results = [
        build_execution_plan(),
        build_revenue_model(),
        build_funnel_design(),
        build_accounting(),
    ]
    print()
    print("=" * 60)
    print("料金モデル v2 (全員同一料金) で再生成完了:")
    print("=" * 60)
    for r in results:
        print(f"  ✓ {r.relative_to(ROOT)}")
    print()
    print("年間売上想定 (保守的):")
    total = sum(p["total"] for p in PROJECTIONS)
    print(f"  ¥{total:,}")
    for p in PROJECTIONS:
        print(f"  {p['phase']} {p['month']:>4}: ¥{p['total']:>12,} (見学会{p['booking_n']}件)")
