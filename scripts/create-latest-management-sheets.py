#!/usr/bin/env python3
"""最新版フォルダ向けの管理シートを生成する。

出力先: docs/最新版/02_管理シート/
  - MVP実行計画.xlsx      (日次/週次/月次のToDoとリリース準備)
  - フェーズ別収益モデル.xlsx (機能x料金x判断基準の対応表)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "docs" / "最新版" / "02_管理シート"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ─── Styling helpers ───────────────────────────────────────────────

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ORANGE_FILL = PatternFill("solid", fgColor="E8740C")
CREAM_FILL = PatternFill("solid", fgColor="FFF8F0")
DARK_FILL = PatternFill("solid", fgColor="3D2200")
LIGHT_GRAY = PatternFill("solid", fgColor="F3F4F6")
GREEN_FILL = PatternFill("solid", fgColor="D1FAE5")
YELLOW_FILL = PatternFill("solid", fgColor="FEF3C7")
RED_FILL = PatternFill("solid", fgColor="FEE2E2")

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
ORANGE_BOLD = Font(bold=True, color="E8740C", size=12)
BOLD = Font(bold=True, size=10)
REG = Font(size=10)
SMALL = Font(size=9, color="6B7280")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def write_header(ws, row: int, headers: list[str], widths: list[int]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = DARK_FILL
        c.font = WHITE_BOLD
        c.alignment = CENTER
        c.border = BORDER
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


def write_row(ws, row: int, values: list, fill: PatternFill | None = None, align=LEFT) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = REG
        c.alignment = align
        c.border = BORDER
        if fill:
            c.fill = fill
    ws.row_dimensions[row].height = 22


def write_title(ws, row: int, title: str, cols: int, subtitle: str | None = None) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=18, color="3D2200")
    c.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 34
    row += 1
    if subtitle:
        c = ws.cell(row=row, column=1, value=subtitle)
        c.font = SMALL
        c.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        row += 1
    return row + 1


def build_execution_plan() -> Path:
    wb = Workbook()

    # ── 0. 目次 ──
    ws = wb.active
    ws.title = "目次"
    row = write_title(
        ws, 1, "ぺいほーむ MVP 実行計画", 3,
        "2026-05-01 リリースに向けた実行管理シート。全てのシートは「誰が・いつまでに・何をするか」で書かれています。",
    )
    write_header(ws, row, ["#", "シート名", "用途"], [6, 32, 60])
    row += 1
    toc = [
        ("①", "リリース前ToDo (4月)", "4/11〜4/30 の作業タスク一覧。担当と期日で管理"),
        ("②", "Phase1 週次ToDo", "5/1〜6/30 の週次運用作業。毎週月曜の朝会で使う"),
        ("③", "Phase別機能追加スケジュール", "Phase 1〜5 で何を、いつ、誰が復活させるか"),
        ("④", "工務店10社 対応管理", "10社それぞれの連絡・契約・撮影・成果の進捗管理"),
        ("⑤", "リリース前チェックリスト", "4/30 までに完了するべき項目の一覧"),
        ("⑥", "月次KPIトラッキング", "リリース後12ヶ月のKPI目標と実績記録欄"),
        ("⑦", "非公開パス復活スケジュール", "40+ 非公開パスの復活タイミング一覧"),
    ]
    for code, name, use in toc:
        write_row(ws, row, [code, name, use])
        row += 1

    # ── ① リリース前ToDo ──
    ws = wb.create_sheet("① リリース前ToDo (4月)")
    row = write_title(ws, 1, "① リリース前ToDo (4月)", 7, "2026-05-01 リリースに向けた確定タスク")
    write_header(
        ws, row,
        ["#", "タスク", "担当", "期日", "所要", "ステータス", "備考"],
        [5, 50, 14, 12, 8, 14, 30],
    )
    row += 1
    release_todos = [
        ("4-1", "Vercel 本番プロジェクト作成", "エンジニア", "2026-04-20", "1h", "未着手", "独自ドメイン payhome.jp を紐付け"),
        ("4-2", "本番用 .env 設定 (NextAuth, Supabase, Stripe)", "エンジニア", "2026-04-20", "2h", "未着手", "NEXTAUTH_SECRET は新規生成"),
        ("4-3", "Supabase 本番マイグレーション適用", "エンジニア", "2026-04-21", "2h", "未着手", "users/leads/events/contact_preferences テーブル"),
        ("4-4", "本番DBに初期データ投入 (工務店10社)", "エンジニア", "2026-04-22", "3h", "未着手", "builders-data.ts と同期"),
        ("4-5", "GA4 計測タグ埋め込み動作確認", "エンジニア", "2026-04-23", "1h", "未着手", "イベント: diagnosis_start, event_booking"),
        ("4-6", "Cloudflare Turnstile 有効化", "エンジニア", "2026-04-23", "2h", "未着手", "signup, contact, event booking"),
        ("4-7", "Sentry エラー監視セットアップ", "エンジニア", "2026-04-24", "2h", "未着手", "Vercel 統合も検討"),
        ("4-8", "17公開ページの E2E スモークテスト", "エンジニア", "2026-04-25", "3h", "未着手", "Playwright で自動化"),
        ("4-9", "40+ 非公開ページの 404 検証", "エンジニア", "2026-04-25", "1h", "未着手", "スクリプトで一括実行"),
        ("4-10", "Search Console 登録 + sitemap 送信", "PM", "2026-04-26", "1h", "未着手", "payhome.jp/sitemap.xml"),
        ("4-11", "工務店10社へリリース予告メール送信", "営業", "2026-04-20", "2h", "未着手", "料金プラン_MVP版.md 添付"),
        ("4-12", "工務店10社と料金プラン最終合意", "営業", "2026-04-27", "−", "未着手", "5/1 以前に書面締結"),
        ("4-13", "営業FAQ の準備・全スタッフ共有", "営業", "2026-04-24", "2h", "未着手", "想定Q&A 20問"),
        ("4-14", "問合せ対応フロー・担当割り当て", "PM", "2026-04-25", "1h", "未着手", "お問い合わせ → 24h以内返信ルール"),
        ("4-15", "経理: 請求書テンプレート作成", "経理", "2026-04-23", "1h", "未着手", "見学会予約 ¥50,000 税込・振込先"),
        ("4-16", "経理: 入金確認フロー設定", "経理", "2026-04-25", "1h", "未着手", "弥生会計 / freee どちらで管理するか決定"),
        ("4-17", "プレスリリース原稿作成", "代表", "2026-04-27", "2h", "未着手", "5/1 AM 9:00 配信想定"),
        ("4-18", "YouTube リリース告知動画 収録・公開", "代表", "2026-04-29", "4h", "未着手", "5/1 AM 6:00 公開"),
        ("4-19", "SNS リリース告知予約投稿 (X, Instagram, Threads)", "代表", "2026-04-30", "1h", "未着手", "5/1 AM 9:00 に自動投稿"),
        ("4-20", "本番切替リハーサル", "エンジニア", "2026-04-29", "2h", "未着手", "Vercel プレビュー → プロダクション昇格の手順確認"),
        ("4-21", "5/1 AM 6:00 本番リリース", "エンジニア", "2026-05-01", "1h", "未着手", "代表と営業もスタンバイ"),
        ("4-22", "リリース後の初日モニタリング", "全員", "2026-05-01", "8h", "未着手", "AM 6:00〜PM 22:00 障害対応体制"),
    ]
    for t in release_todos:
        write_row(ws, row, list(t))
        row += 1
    ws.freeze_panes = "A3"

    # ── ② Phase1 週次ToDo ──
    ws = wb.create_sheet("② Phase1 週次ToDo")
    row = write_title(ws, 1, "② Phase 1 週次運用ToDo (5/1〜6/30)", 6, "毎週月曜 9:00 の朝会で確認")
    write_header(
        ws, row,
        ["曜日", "時刻", "タスク", "担当", "頻度", "判断基準"],
        [8, 10, 45, 14, 10, 40],
    )
    row += 1
    weekly = [
        ("月", "9:00", "前週の見学会予約・リード・診断数の集計", "PM", "毎週", "先週比+10%以上なら維持、-10%なら原因分析"),
        ("月", "9:30", "エラーログの確認 (Sentry / Vercel)", "エンジニア", "毎週", "Critical 0件を維持"),
        ("月", "10:00", "工務店10社への週次レポート送信", "営業", "毎週", "全社送信完了で✓"),
        ("月", "15:00", "YouTube 来週の動画スクリプト確認", "代表", "毎週", "−"),
        ("火", "10:00", "問合せフォームから来た連絡への対応", "営業", "毎日", "24h以内に全件返信"),
        ("水", "10:00", "YouTube 新着動画の公開 + ぺいほーむ掲載", "代表", "毎週水", "YouTube公開から1時間以内に掲載"),
        ("水", "14:00", "SNS: 動画公開の拡散 (X, Instagram, Threads)", "代表", "毎週水", "全プラットフォーム投稿で✓"),
        ("金", "10:00", "見学会予約の工務店確認 (通知届いたか)", "営業", "毎週", "全件確認で✓"),
        ("金", "15:00", "工務店10社のうち1社と電話フォローアップ", "営業", "毎週", "週1社 × 10週 = 全社巡回"),
        ("金", "17:00", "週次 KPI を経理・代表に共有", "PM", "毎週", "スプレッドシート更新"),
        ("土", "−", "深刻な障害発生時のオンコール", "エンジニア", "随時", "−"),
        ("日", "−", "ブログ/Note の記事執筆 (月2本目標)", "代表", "隔週", "−"),
    ]
    for item in weekly:
        write_row(ws, row, list(item))
        row += 1
    ws.freeze_panes = "A3"

    # ── ③ Phase別機能追加スケジュール ──
    ws = wb.create_sheet("③ Phase別機能追加")
    row = write_title(ws, 1, "③ Phase別 機能追加スケジュール", 7, "いつ何を復活させるか。エンジニアの実装計画と同期")
    write_header(
        ws, row,
        ["Phase", "期間", "追加機能", "対応パス", "担当", "期日", "収益モデルへの影響"],
        [8, 20, 26, 32, 14, 14, 38],
    )
    row += 1
    phase_features = [
        ("P1", "2026/05/01〜06/30", "運用のみ", "−", "運用チーム", "2026-06-30", "見学会予約¥50,000 + 3% を運用開始"),
        ("P2", "2026/07/01〜08/31", "業界ニュース", "/news, /news/[id]", "エンジニア", "2026-07-01", "−"),
        ("P2", "2026/07/01〜08/31", "記事・コラム", "/articles, /articles/[id]", "エンジニア", "2026-07-01", "−"),
        ("P2", "2026/07/01〜08/31", "管理画面 ニュース/記事CMS", "/admin/news, /admin/articles", "エンジニア", "2026-07-03", "−"),
        ("P2", "2026/07/01〜08/31", "B2B ニュース", "/biz/news", "エンジニア", "2026-07-15", "−"),
        ("P2", "2026/07/01〜08/31", "撮影プラン申込フォーム", "内部実装", "エンジニア", "2026-08-15", "撮影費 ¥150,000/本 開始"),
        ("P3", "2026/09/01〜11/30", "販売中の分譲戸建", "/sale-homes, /sale-homes/[id]", "エンジニア", "2026-09-01", "月額掲載 ¥30,000 開始"),
        ("P3", "2026/09/01〜11/30", "取扱中の土地", "/lands, /lands/[id]", "エンジニア", "2026-09-01", "月額掲載 ¥30,000 開始"),
        ("P3", "2026/09/01〜11/30", "特集", "/features, /features/[id]", "エンジニア", "2026-09-01", "−"),
        ("P3", "2026/09/01〜11/30", "ローンシミュレーター", "/simulator", "エンジニア", "2026-09-05", "−"),
        ("P3", "2026/09/01〜11/30", "エリア別ページ", "/area", "エンジニア", "2026-09-10", "−"),
        ("P3", "2026/09/01〜11/30", "工務店 建売/土地CMS", "/dashboard/builder/sale-homes, /lands", "エンジニア", "2026-09-01", "−"),
        ("P3", "2026/09/01〜11/30", "工務店比較", "/builders/compare", "エンジニア", "2026-10-01", "−"),
        ("P4", "2026/12/01〜2027/02/28", "匿名AI仲介質問", "/mypage/questions, /dashboard/builder/questions", "エンジニア", "2026-12-01", "SaaS ¥50,000/月 開始"),
        ("P4", "2026/12/01〜2027/02/28", "ウェビナー", "/webinar, /webinar/[id]", "エンジニア", "2026-12-15", "SaaS に含む"),
        ("P4", "2026/12/01〜2027/02/28", "マガジン", "/magazine", "エンジニア", "2027-01-15", "SaaS に含む"),
        ("P4", "2026/12/01〜2027/02/28", "インタビュー", "/interview, /interview/[id]", "エンジニア", "2027-01-15", "SaaS に含む"),
        ("P4", "2026/12/01〜2027/02/28", "Stripe決済 + 請求自動化", "内部実装", "エンジニア", "2026-11-30", "全プランに必須"),
        ("P4", "2026/12/01〜2027/02/28", "管理画面 全CMS", "/admin/* 全て", "エンジニア", "2027-01-31", "−"),
        ("P5", "2027/03〜", "全国工務店追加", "builders-data 拡張", "営業+エンジニア", "2027-03-31", "提携料は正規価格"),
    ]
    for f in phase_features:
        write_row(ws, row, list(f))
        row += 1
    ws.freeze_panes = "A3"

    # ── ④ 工務店10社 対応管理 ──
    ws = wb.create_sheet("④ 工務店10社 対応管理")
    row = write_title(ws, 1, "④ 工務店10社 対応管理", 11, "契約・撮影・成果の進捗を社単位で追跡")
    write_header(
        ws, row,
        [
            "#",
            "社名",
            "担当",
            "契約書",
            "MVP説明",
            "撮影予定",
            "掲載URL",
            "累計予約数",
            "累計成果報酬",
            "最終連絡日",
            "次回アクション",
        ],
        [4, 22, 12, 10, 10, 12, 28, 10, 12, 12, 30],
    )
    row += 1
    # Placeholder rows — builders-data.ts と同期して営業が埋める
    for i in range(1, 11):
        write_row(
            ws, row,
            [i, f"(社名 #{i})", "", "未締結", "未実施", "未定", "", 0, 0, "", ""],
        )
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑤ リリース前チェックリスト ──
    ws = wb.create_sheet("⑤ リリース前チェックリスト")
    row = write_title(ws, 1, "⑤ リリース前チェックリスト", 5, "4/30 までに ✓ を全て埋める")
    write_header(ws, row, ["#", "カテゴリ", "項目", "責任者", "完了 ✓"], [5, 14, 55, 14, 10])
    row += 1
    checklist = [
        ("1", "インフラ", "Vercel 本番プロジェクト作成完了", "エンジニア", ""),
        ("2", "インフラ", "独自ドメイン payhome.jp SSL 有効", "エンジニア", ""),
        ("3", "インフラ", "Supabase 本番環境マイグレーション完了", "エンジニア", ""),
        ("4", "インフラ", "本番 .env 全キー設定済み", "エンジニア", ""),
        ("5", "セキュリティ", "NEXTAUTH_SECRET 本番用を新規生成", "エンジニア", ""),
        ("6", "セキュリティ", "Turnstile (bot対策) 有効", "エンジニア", ""),
        ("7", "セキュリティ", "管理画面 admin ロール認証動作確認", "エンジニア", ""),
        ("8", "セキュリティ", "工務店ダッシュボード builder ロール確認", "エンジニア", ""),
        ("9", "計測", "GA4 計測タグ動作確認", "エンジニア", ""),
        ("10", "計測", "Sentry エラー監視セットアップ", "エンジニア", ""),
        ("11", "計測", "Vercel Analytics 有効", "エンジニア", ""),
        ("12", "テスト", "17公開ページ 200 確認", "エンジニア", ""),
        ("13", "テスト", "40+非公開ページ 404 確認", "エンジニア", ""),
        ("14", "テスト", "AI診断フロー完走確認", "エンジニア", ""),
        ("15", "テスト", "見学会予約フロー完走確認", "エンジニア", ""),
        ("16", "テスト", "会員登録・ログイン動作確認", "エンジニア", ""),
        ("17", "テスト", "モバイル表示 (iPhone, Android) 確認", "エンジニア", ""),
        ("18", "SEO", "robots.txt 本番URLで閲覧可能", "エンジニア", ""),
        ("19", "SEO", "sitemap.xml 本番URLで閲覧可能", "エンジニア", ""),
        ("20", "SEO", "Search Console 登録 + sitemap 送信", "PM", ""),
        ("21", "SEO", "OGP 画像 + Twitter Card 動作確認", "エンジニア", ""),
        ("22", "営業", "10社と料金プラン書面締結", "営業", ""),
        ("23", "営業", "10社の掲載情報最終チェック", "営業", ""),
        ("24", "営業", "営業FAQ 20問 全スタッフ共有", "営業", ""),
        ("25", "営業", "リリース予告メール10社送信", "営業", ""),
        ("26", "オペレーション", "問合せ対応フロー決定", "PM", ""),
        ("27", "オペレーション", "見学会予約 → 工務店通知ルート動作確認", "PM", ""),
        ("28", "経理", "請求書テンプレ作成", "経理", ""),
        ("29", "経理", "振込先確認・社内共有", "経理", ""),
        ("30", "広報", "プレスリリース原稿承認", "代表", ""),
        ("31", "広報", "YouTube 告知動画 公開予約", "代表", ""),
        ("32", "広報", "X/Instagram/Threads 告知予約", "代表", ""),
        ("33", "リリース", "リリースリハーサル完了", "エンジニア", ""),
        ("34", "リリース", "ロールバック手順確認", "エンジニア", ""),
    ]
    for r in checklist:
        write_row(ws, row, list(r))
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑥ 月次KPIトラッキング ──
    ws = wb.create_sheet("⑥ 月次KPI")
    row = write_title(ws, 1, "⑥ 月次 KPI トラッキング", 14, "Phase 1〜5 の月次KPI目標。実績は毎月末に記入")
    headers = ["月", "Phase"] + [f"{m}月" for m in [5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3, 4]]
    write_header(ws, row, headers, [18, 6] + [9] * 12)
    row += 1

    kpi_rows = [
        ("月間UU", ""),
        ("目標 UU", "",),
        ("実績 UU", ""),
        ("AI診断 完了数", ""),
        ("目標", ""),
        ("実績", ""),
        ("見学会予約 数", ""),
        ("目標", ""),
        ("実績", ""),
        ("撮影プラン 本数", ""),
        ("目標", ""),
        ("実績", ""),
        ("月額掲載 社数", ""),
        ("目標", ""),
        ("実績", ""),
        ("SaaS 社数", ""),
        ("目標", ""),
        ("実績", ""),
        ("月次売上 (円)", ""),
        ("目標", ""),
        ("実績", ""),
    ]
    targets = {
        "月間UU": [5000, 7500, 10000, 12500, 15000, 17500, 20000, 25000, 30000, 33000, 36000, 40000],
        "AI診断 完了数": [200, 300, 400, 500, 600, 700, 800, 1000, 1200, 1300, 1400, 1500],
        "見学会予約 数": [10, 15, 20, 25, 30, 35, 40, 50, 55, 60, 65, 70],
        "撮影プラン 本数": [0, 0, 3, 5, 5, 6, 6, 7, 7, 8, 8, 9],
        "月額掲載 社数": [0, 0, 0, 0, 3, 5, 6, 8, 9, 10, 11, 12],
        "SaaS 社数": [0, 0, 0, 0, 0, 0, 0, 2, 4, 6, 8, 10],
        "月次売上 (円)": [
            150_000, 250_000, 400_000, 550_000, 750_000, 950_000, 1_100_000,
            1_400_000, 1_650_000, 1_850_000, 2_100_000, 2_350_000,
        ],
    }
    phases_by_month = ["P1", "P1", "P2", "P2", "P3", "P3", "P3", "P4", "P4", "P4", "P5", "P5"]
    for label, _ in kpi_rows:
        if label in targets:
            values = [label, ""] + targets[label]
            write_row(ws, row, values, fill=CREAM_FILL)
        elif label == "目標":
            write_row(ws, row, [label, ""] + [""] * 12, fill=YELLOW_FILL)
        elif label == "実績":
            write_row(ws, row, [label, ""] + [""] * 12, fill=GREEN_FILL)
        else:
            write_row(ws, row, [label, ""] + [""] * 12)
        row += 1
    # Phase row
    row_phase = 3
    ws.cell(row=row_phase, column=2, value="").fill = ORANGE_FILL
    for i, p in enumerate(phases_by_month, start=3):
        ws.cell(row=row_phase, column=i, value=p).fill = CREAM_FILL

    ws.freeze_panes = "C3"

    # ── ⑦ 非公開パス復活スケジュール ──
    ws = wb.create_sheet("⑦ 非公開パス復活")
    row = write_title(ws, 1, "⑦ 非公開パス復活スケジュール", 6, "middleware.ts の HIDDEN_PATH_PREFIXES から削除するタイミング")
    write_header(
        ws, row,
        ["#", "パス", "Phase", "復活予定", "実装担当", "備考"],
        [5, 40, 8, 14, 14, 38],
    )
    row += 1
    restore = [
        (1, "/news", "P2", "2026-07-01", "エンジニア", "業界ニュース一覧・詳細"),
        (2, "/articles", "P2", "2026-07-01", "エンジニア", "記事・コラム"),
        (3, "/admin/news", "P2", "2026-07-03", "エンジニア", "ニュースCMS"),
        (4, "/admin/articles", "P2", "2026-07-03", "エンジニア", "記事CMS"),
        (5, "/biz/news", "P2", "2026-07-15", "エンジニア", "B2B業界ニュース"),
        (6, "/biz/articles", "P2", "2026-08-01", "エンジニア", "B2B集客ノウハウ"),
        (7, "/sale-homes", "P3", "2026-09-01", "エンジニア", "建売一覧・詳細"),
        (8, "/lands", "P3", "2026-09-01", "エンジニア", "土地一覧・詳細"),
        (9, "/features", "P3", "2026-09-01", "エンジニア", "特集"),
        (10, "/simulator", "P3", "2026-09-05", "エンジニア", "ローンシミュレーター"),
        (11, "/area", "P3", "2026-09-10", "エンジニア", "エリア別ページ"),
        (12, "/dashboard/builder/sale-homes", "P3", "2026-09-01", "エンジニア", "工務店 建売CMS"),
        (13, "/dashboard/builder/lands", "P3", "2026-09-01", "エンジニア", "工務店 土地CMS"),
        (14, "/admin/sale-homes", "P3", "2026-09-01", "エンジニア", "管理 建売CMS"),
        (15, "/admin/lands", "P3", "2026-09-01", "エンジニア", "管理 土地CMS"),
        (16, "/admin/features", "P3", "2026-09-01", "エンジニア", "管理 特集CMS"),
        (17, "/builders/compare", "P3", "2026-10-01", "エンジニア", "工務店比較ページ"),
        (18, "/mypage/questions", "P4", "2026-12-01", "エンジニア", "匿名AI仲介質問"),
        (19, "/dashboard/builder/questions", "P4", "2026-12-01", "エンジニア", "工務店Q&A"),
        (20, "/admin/reports", "P4", "2026-12-01", "エンジニア", "月次レポート自動集計"),
        (21, "/webinar", "P4", "2026-12-15", "エンジニア", "ウェビナー一覧・詳細"),
        (22, "/interview", "P4", "2027-01-15", "エンジニア", "工務店インタビュー"),
        (23, "/magazine", "P4", "2027-01-15", "エンジニア", "マガジン"),
        (24, "/biz/webinar", "P4", "2026-12-15", "エンジニア", "B2Bウェビナー"),
        (25, "/dashboard/builder/billing", "P4", "2026-12-01", "エンジニア", "工務店 請求・プラン管理 (Stripe)"),
        (26, "/biz/service", "P4", "2026-12-01", "エンジニア", "B2Bサービス概要 (大幅リニューアル)"),
        (27, "/biz/ad", "P4", "2026-12-01", "エンジニア", "B2B広告・タイアップ"),
        (28, "/biz/partner", "P4", "2026-12-01", "エンジニア", "B2B提携パートナー募集"),
        (29, "/dashboard/builder/case-studies", "P4", "2027-01-01", "エンジニア", "工務店 施工事例CMS"),
        (30, "/dashboard/builder/videos", "P4", "2027-01-01", "エンジニア", "工務店 動画CMS"),
        (31, "/dashboard/builder/reviews", "P4", "2027-01-01", "エンジニア", "工務店 お客様の声CMS"),
        (32, "/admin/interviews", "P4", "2027-01-15", "エンジニア", "管理 インタビューCMS"),
        (33, "/admin/webinars", "P4", "2026-12-15", "エンジニア", "管理 ウェビナーCMS"),
        (34, "/admin/magazine", "P4", "2027-01-15", "エンジニア", "管理 マガジンCMS"),
        (35, "/admin/videos", "P4", "2027-01-01", "エンジニア", "管理 動画CMS"),
        (36, "/admin/case-studies", "P4", "2027-01-01", "エンジニア", "管理 施工事例CMS"),
        (37, "/admin/reviews", "P4", "2027-01-01", "エンジニア", "管理 お客様の声CMS"),
        (38, "/admin/biz-articles", "P4", "2027-01-15", "エンジニア", "管理 B2B記事CMS"),
        (39, "/admin/biz-news", "P4", "2027-01-15", "エンジニア", "管理 B2BニュースCMS"),
        (40, "/admin/biz-webinars", "P4", "2027-01-15", "エンジニア", "管理 B2BウェビナーCMS"),
        (41, "/admin/users", "P4", "2027-02-01", "エンジニア", "ユーザー管理"),
        (42, "/admin/properties", "P4", "2027-02-01", "エンジニア", "物件データ管理"),
        (43, "/admin/security, /activity, /system, /notifications, /data", "P4", "2027-02-15", "エンジニア", "システム運用系画面"),
        (44, "/mypage/catalog", "P4", "2027-01-01", "エンジニア", "マイページのカタログダウンロード"),
        (45, "/mypage/feedback", "P4", "2027-02-01", "エンジニア", "マイページ フィードバック"),
        (46, "/welcome, /property, /builders/contact", "P5", "2027-03-01", "エンジニア", "全国展開時に再検討"),
    ]
    for r in restore:
        write_row(ws, row, list(r))
        row += 1
    ws.freeze_panes = "A3"

    out = OUT_DIR / "MVP実行計画.xlsx"
    wb.save(out)
    return out


def build_revenue_model() -> Path:
    wb = Workbook()

    # ── 目次 ──
    ws = wb.active
    ws.title = "目次"
    row = write_title(
        ws, 1, "ぺいほーむ フェーズ別 収益モデル", 3,
        "Phase x 機能 x 収益モデル x 10社優遇 の対応表",
    )
    write_header(ws, row, ["#", "シート名", "用途"], [6, 30, 60])
    row += 1
    toc = [
        ("①", "料金プラン一覧", "Phase 1〜5 の料金プランと10社優遇"),
        ("②", "機能x料金マトリクス", "どの機能がどの料金プランで使えるか"),
        ("③", "月次売上シミュレーション", "12ヶ月の想定売上と内訳"),
        ("④", "工務店別プラン管理", "10社 + 新規それぞれの適用プラン"),
        ("⑤", "Phase移行 判断基準", "次のPhaseへ進む条件"),
        ("⑥", "プラン別の原価", "見学会予約/撮影/掲載プランの原価と粗利"),
    ]
    for t in toc:
        write_row(ws, row, list(t))
        row += 1

    # ── ① 料金プラン一覧 ──
    ws = wb.create_sheet("① 料金プラン一覧")
    row = write_title(ws, 1, "① Phase別 料金プラン一覧", 6, "優遇は既存10社に対して常に適用")
    write_header(
        ws, row,
        ["Phase", "項目", "正規価格", "10社優遇", "課金タイミング", "備考"],
        [8, 28, 16, 18, 20, 32],
    )
    row += 1
    plans = [
        ("P1", "見学会予約", "¥50,000/件", "初月2ヶ月 無料", "予約成立時", "5/1-6/30 は運用検証のため無料提供"),
        ("P1", "成約成果報酬", "3% (上限 ¥500万)", "1.5% (上限 ¥250万)", "契約成立時", "建物本体価格の3%"),
        ("P1", "AI家づくり診断", "無料", "無料", "−", "ユーザー無料"),
        ("P2", "見学会予約", "¥50,000/件", "¥25,000/件", "予約成立時", "以降P2〜P5 常に半額"),
        ("P2", "成約成果報酬", "3%", "1.5%", "契約成立時", "以降P2〜P5 常に半額"),
        ("P2", "撮影プラン", "¥150,000/本", "¥75,000/本", "撮影日", "ルームツアー撮影 + 編集 + YouTube 掲載"),
        ("P3", "見学会予約", "¥50,000/件", "¥25,000/件", "予約成立時", "継続"),
        ("P3", "成約成果報酬", "3%", "1.5%", "契約成立時", "継続"),
        ("P3", "撮影プラン", "¥150,000/本", "¥75,000/本", "撮影日", "継続"),
        ("P3", "月額掲載 ベーシック", "¥30,000/月", "初年度 無料", "毎月月末請求", "物件3件まで掲載"),
        ("P3", "月額掲載 プロ", "¥80,000/月", "初年度 ¥40,000/月", "毎月月末請求", "物件無制限 + 月1本特集枠"),
        ("P4", "見学会予約", "¥50,000/件", "¥25,000/件", "予約成立時", "継続"),
        ("P4", "成約成果報酬", "3%", "1.5%", "契約成立時", "継続"),
        ("P4", "撮影プラン", "¥150,000/本", "¥75,000/本", "撮影日", "継続"),
        ("P4", "月額掲載 ベーシック", "¥30,000/月", "¥15,000/月", "毎月月末請求", "10社は優遇継続"),
        ("P4", "月額掲載 プロ", "¥80,000/月", "¥40,000/月", "毎月月末請求", "継続"),
        ("P4", "SaaS ライト", "¥50,000/月", "¥25,000/月", "毎月月末請求", "AI質問無制限 + インタビュー掲載 + 月次レポート"),
        ("P4", "SaaS プロ", "¥150,000/月", "¥75,000/月", "毎月月末請求", "全機能 + 専任担当 + API 連携"),
        ("P5", "見学会予約", "¥50,000/件", "¥25,000/件", "予約成立時", "継続"),
        ("P5", "成約成果報酬", "3%", "1.5%", "契約成立時", "継続"),
        ("P5", "撮影プラン", "¥150,000/本", "¥75,000/本", "撮影日", "継続"),
        ("P5", "月額掲載 ベーシック", "¥30,000/月", "¥15,000/月", "毎月月末請求", "永続優遇"),
        ("P5", "月額掲載 プロ", "¥80,000/月", "¥40,000/月", "毎月月末請求", "永続優遇"),
        ("P5", "SaaS ライト", "¥50,000/月", "¥25,000/月", "毎月月末請求", "永続優遇"),
        ("P5", "SaaS プロ", "¥150,000/月", "¥75,000/月", "毎月月末請求", "永続優遇"),
    ]
    for p in plans:
        write_row(ws, row, list(p))
        row += 1
    ws.freeze_panes = "A3"

    # ── ② 機能x料金マトリクス ──
    ws = wb.create_sheet("② 機能x料金")
    row = write_title(ws, 1, "② 機能 x 料金プラン マトリクス", 7, "どの機能がどの料金プランで提供されるか")
    write_header(
        ws, row,
        ["機能", "無料", "見学会予約", "撮影プラン", "月額 ベーシック", "月額 プロ", "SaaS"],
        [30, 10, 12, 12, 16, 12, 10],
    )
    row += 1
    matrix = [
        ("ぺいほーむ掲載 (基本情報)", "✓", "✓", "✓", "✓", "✓", "✓"),
        ("AI家づくり診断 送客", "✓", "✓", "✓", "✓", "✓", "✓"),
        ("見学会予約 受付", "−", "✓", "✓", "✓", "✓", "✓"),
        ("ルームツアー動画 掲載", "−", "−", "✓ (撮影時)", "−", "−", "−"),
        ("物件 3件掲載", "−", "−", "−", "✓", "✓", "✓"),
        ("物件 無制限掲載", "−", "−", "−", "−", "✓", "✓"),
        ("特集枠 月1本", "−", "−", "−", "−", "✓", "✓"),
        ("AI質問 対応 (無制限)", "−", "−", "−", "−", "−", "✓"),
        ("インタビュー 掲載", "−", "−", "−", "−", "−", "✓"),
        ("月次レポート 自動配信", "−", "−", "−", "−", "−", "✓"),
        ("API 連携", "−", "−", "−", "−", "−", "Pro のみ"),
        ("専任担当サポート", "−", "−", "−", "−", "−", "Pro のみ"),
    ]
    for m in matrix:
        write_row(ws, row, list(m))
        row += 1
    ws.freeze_panes = "A3"

    # ── ③ 月次売上シミュレーション ──
    ws = wb.create_sheet("③ 月次売上シミュレーション")
    row = write_title(
        ws, 1, "③ 月次売上シミュレーション (12ヶ月)", 8,
        "10社優遇 + 新規11社目以降を含む保守的な想定。Phase2以降は新規開拓により上振れ可能",
    )
    write_header(
        ws, row,
        ["月", "Phase", "見学会予約", "成果報酬", "撮影", "月額掲載", "SaaS", "月次合計"],
        [8, 6, 14, 14, 14, 14, 14, 16],
    )
    row += 1

    # Simulation data (conservative)
    sim = [
        # month, phase, booking, success, filming, subscription, saas
        ("5月", "P1", 0, 0, 0, 0, 0),                        # 優遇で無料
        ("6月", "P1", 50_000, 125_000, 0, 0, 0),             # 新規1件¥50k + 成果1件¥125k
        ("7月", "P2", 125_000, 125_000, 75_000, 0, 0),       # 10社半額 + 新規1件 + 撮影1本
        ("7月(補正)", "", "", "", "", "", "", ""),  # placeholder for visibility, removed below
        ("8月", "P2", 175_000, 125_000, 150_000, 0, 0),
        ("9月", "P3", 225_000, 250_000, 225_000, 30_000, 0),  # 新規1社月額加入
        ("10月", "P3", 275_000, 250_000, 225_000, 90_000, 0),
        ("11月", "P3", 325_000, 375_000, 300_000, 180_000, 0),
        ("12月", "P4", 375_000, 500_000, 300_000, 240_000, 75_000),
        ("1月", "P4", 425_000, 500_000, 375_000, 270_000, 150_000),
        ("2月", "P4", 475_000, 625_000, 375_000, 300_000, 225_000),
        ("3月", "P5", 500_000, 750_000, 450_000, 360_000, 300_000),
        ("4月", "P5", 550_000, 750_000, 450_000, 390_000, 375_000),
    ]
    # Filter out placeholder rows
    sim = [s for s in sim if s[0] != "7月(補正)"]
    for month, phase, booking, success, filming, sub, saas in sim:
        total = booking + success + filming + sub + saas
        values = [
            month,
            phase,
            f"¥{booking:,}",
            f"¥{success:,}",
            f"¥{filming:,}",
            f"¥{sub:,}",
            f"¥{saas:,}",
            f"¥{total:,}",
        ]
        fill = LIGHT_GRAY if total < 500_000 else (CREAM_FILL if total < 1_500_000 else GREEN_FILL)
        write_row(ws, row, values, fill=fill, align=CENTER)
        row += 1
    # Totals
    total_year = sum(
        booking + success + filming + sub + saas
        for _, _, booking, success, filming, sub, saas in sim
    )
    write_row(
        ws, row,
        ["年間合計", "", "", "", "", "", "", f"¥{total_year:,}"],
        fill=ORANGE_FILL,
        align=CENTER,
    )
    ws.cell(row=row, column=1).font = WHITE_BOLD
    ws.cell(row=row, column=8).font = WHITE_BOLD

    ws.freeze_panes = "A3"

    # ── ④ 工務店別プラン管理 ──
    ws = wb.create_sheet("④ 工務店別プラン管理")
    row = write_title(ws, 1, "④ 工務店別プラン管理", 9, "社ごとの契約プラン・優遇・期間を管理")
    write_header(
        ws, row,
        ["#", "社名", "契約日", "分類", "見学会予約", "成果報酬", "撮影プラン", "月額掲載", "SaaS"],
        [4, 22, 12, 14, 14, 12, 14, 18, 14],
    )
    row += 1
    for i in range(1, 11):
        write_row(
            ws, row,
            [
                i,
                f"(既存10社 #{i})",
                "2026-04-XX",
                "既存優遇",
                "¥25,000/件 (P2〜)",
                "1.5%",
                "¥75,000/本 (P2〜)",
                "初年度無料 (P3〜)",
                "¥25,000/月 (P4〜)",
            ],
            fill=CREAM_FILL,
        )
        row += 1
    for i in range(11, 16):
        write_row(
            ws, row,
            [
                i,
                "(新規 ここから)",
                "",
                "新規",
                "¥50,000/件",
                "3%",
                "¥150,000/本",
                "¥30,000/月",
                "¥50,000/月",
            ],
        )
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑤ Phase移行 判断基準 ──
    ws = wb.create_sheet("⑤ Phase移行 判断基準")
    row = write_title(ws, 1, "⑤ Phase 移行 判断基準", 4, "次のフェーズに進むかどうかを決めるKPI")
    write_header(ws, row, ["Phase", "判断タイミング", "必須条件", "補足条件"], [8, 18, 55, 50])
    row += 1
    criteria = [
        ("P0→P1", "2026-04-30", "17公開ページ200・40+非公開404・10社と書面締結", "プレスリリース準備完了"),
        ("P1→P2", "2026-06-30", "見学会予約が2ヶ月連続で月5件以上", "10社のうち8社がPhase2プランに興味あり"),
        ("P2→P3", "2026-08-31", "撮影プラン申込が2ヶ月で5本以上、見学会予約が月15件以上", "記事/ニュースCMSが安定稼働"),
        ("P3→P4", "2026-11-30", "月額掲載の新規契約3社以上、総売上月50万円超", "物件情報の登録が工務店から安定"),
        ("P4→P5", "2027-02-28", "SaaSプラン契約10社以上、総売上月200万円超", "全国展開の準備完了 (マーケ・採用計画)"),
    ]
    for c in criteria:
        write_row(ws, row, list(c), align=LEFT_TOP)
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑥ プラン別の原価 ──
    ws = wb.create_sheet("⑥ プラン別原価")
    row = write_title(ws, 1, "⑥ プラン別 原価と粗利", 5, "プラットフォームの採算性評価")
    write_header(ws, row, ["プラン", "正規価格", "変動原価", "粗利", "粗利率"], [30, 18, 18, 18, 12])
    row += 1
    cost = [
        ("見学会予約 (件あたり)", 50_000, 5_000, 45_000, "90%"),     # 決済手数料・通知コスト
        ("成果報酬 3% (¥3000万契約想定)", 900_000, 20_000, 880_000, "98%"),
        ("撮影プラン (本あたり)", 150_000, 80_000, 70_000, "47%"),   # カメラマン・編集・機材
        ("月額掲載 ベーシック", 30_000, 3_000, 27_000, "90%"),
        ("月額掲載 プロ", 80_000, 10_000, 70_000, "88%"),
        ("SaaS ライト", 50_000, 8_000, 42_000, "84%"),               # AI API コスト想定
        ("SaaS プロ", 150_000, 30_000, 120_000, "80%"),              # 専任担当コスト
    ]
    for name, price, var, profit, rate in cost:
        write_row(
            ws, row,
            [name, f"¥{price:,}", f"¥{var:,}", f"¥{profit:,}", rate],
            align=CENTER,
        )
        row += 1
    ws.freeze_panes = "A3"

    out = OUT_DIR / "フェーズ別収益モデル.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    p1 = build_execution_plan()
    p2 = build_revenue_model()
    print("Created:", p1)
    print("Created:", p2)
