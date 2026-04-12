#!/usr/bin/env python3
"""最新版の経理系資料と動線設計シートを生成する。

出力先:
  docs/最新版/04_経理関係/請求・P&L管理.xlsx
  docs/最新版/02_管理シート/動線設計シート.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_KEIRI = ROOT / "docs" / "最新版" / "04_経理関係"
OUT_MGMT = ROOT / "docs" / "最新版" / "02_管理シート"
OUT_KEIRI.mkdir(parents=True, exist_ok=True)
OUT_MGMT.mkdir(parents=True, exist_ok=True)

# ─── Styling helpers ───────────────────────────────────────────────

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ORANGE = PatternFill("solid", fgColor="E8740C")
CREAM = PatternFill("solid", fgColor="FFF8F0")
DARK = PatternFill("solid", fgColor="3D2200")
GRAY = PatternFill("solid", fgColor="F3F4F6")
GREEN = PatternFill("solid", fgColor="D1FAE5")
YELLOW = PatternFill("solid", fgColor="FEF3C7")
RED = PatternFill("solid", fgColor="FEE2E2")
BLUE = PatternFill("solid", fgColor="DBEAFE")

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True, size=10)
REG = Font(size=10)
SMALL = Font(size=9, color="6B7280")
TITLE_FONT = Font(bold=True, size=18, color="3D2200")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def H(ws, row: int, headers: list[str], widths: list[int]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = DARK
        c.font = WHITE_BOLD
        c.alignment = CENTER
        c.border = BORDER
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


def R(ws, row: int, values: list, fill=None, align=LEFT, bold=False) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(bold=bold, size=10) if bold else REG
        c.alignment = align
        c.border = BORDER
        if fill:
            c.fill = fill
    ws.row_dimensions[row].height = 22


def T(ws, row: int, title: str, cols: int, subtitle: str | None = None) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 34
    row += 1
    if subtitle:
        c = ws.cell(row=row, column=1, value=subtitle)
        c.font = SMALL
        c.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        row += 1
    return row + 1


# ══════════════════════════════════════════════════════════════════
# 1. 経理関係: 請求・P&L 管理
# ══════════════════════════════════════════════════════════════════

def build_accounting() -> Path:
    wb = Workbook()

    # ── 目次 ──
    ws = wb.active
    ws.title = "目次"
    row = T(
        ws, 1, "ぺいほーむ 請求・P&L 管理", 3,
        "2026-05-01 MVPリリース以降の経理管理。各シートの更新タイミングを守って月次で回す",
    )
    H(ws, row, ["#", "シート名", "用途"], [6, 32, 55])
    row += 1
    toc = [
        ("①", "月次請求台帳", "毎月の請求明細 (工務店 x 項目 x 金額)"),
        ("②", "入金管理", "請求に対する入金確認"),
        ("③", "月次P&L 12ヶ月", "Phase 1〜5 の月次P&L想定"),
        ("④", "原価・変動費", "プラン別の原価と粗利率"),
        ("⑤", "固定費 年間計画", "人件費・SaaS・家賃などの固定費"),
        ("⑥", "税務・請求フロー", "請求書発行から入金確認までの手順"),
    ]
    for t in toc:
        R(ws, row, list(t))
        row += 1

    # ── ① 月次請求台帳 ──
    ws = wb.create_sheet("① 月次請求台帳")
    row = T(ws, 1, "① 月次請求台帳", 9, "毎月月末に当月分の請求を記載")
    H(
        ws, row,
        ["請求月", "請求日", "請求先", "項目", "単価", "数量", "小計", "税別/込", "ステータス"],
        [10, 12, 22, 20, 12, 8, 14, 10, 12],
    )
    row += 1
    # Sample template rows
    samples = [
        ("2026-05", "", "(10社のうちのA社)", "成果報酬 1.5%", "", 0, 0, "税別", "未請求"),
        ("2026-05", "", "(10社のうちのB社)", "成果報酬 1.5%", "", 0, 0, "税別", "未請求"),
        ("2026-06", "", "", "見学会予約 (10社無料)", "", 0, 0, "税別", "無料"),
        ("2026-07", "", "", "見学会予約 ¥25,000/件", 25000, 0, 0, "税別", "未請求"),
        ("2026-07", "", "", "撮影プラン ¥75,000/本", 75000, 0, 0, "税別", "未請求"),
        ("2026-09", "", "", "月額掲載ベーシック(10社初年度無料)", 0, 0, 0, "−", "無料"),
        ("2026-09", "", "", "月額掲載ベーシック(新規)", 30000, 0, 0, "税別", "未請求"),
        ("2026-12", "", "", "SaaSライト ¥25,000/月 (10社)", 25000, 0, 0, "税別", "未請求"),
    ]
    for s in samples:
        R(ws, row, list(s), fill=GRAY if s[-1] != "未請求" else None)
        row += 1
    # Add 20 empty rows for data entry
    for _ in range(20):
        R(ws, row, [""] * 9)
        row += 1
    ws.freeze_panes = "A3"

    # ── ② 入金管理 ──
    ws = wb.create_sheet("② 入金管理")
    row = T(ws, 1, "② 入金管理", 8, "請求に対する入金の確認")
    H(
        ws, row,
        ["請求月", "請求先", "項目", "請求額", "支払期限", "入金日", "入金額", "差額"],
        [10, 22, 20, 14, 12, 12, 14, 12],
    )
    row += 1
    for _ in range(30):
        R(ws, row, [""] * 8)
        row += 1
    ws.freeze_panes = "A3"

    # ── ③ 月次P&L 12ヶ月 ──
    ws = wb.create_sheet("③ 月次P&L")
    row = T(ws, 1, "③ 月次 P&L (12ヶ月計画)", 14, "フェーズ別収益モデル.xlsx と連動。毎月実績を記入")
    H(
        ws, row,
        ["項目", "Phase"] + [f"{m}月" for m in [5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3, 4]],
        [24, 6] + [11] * 12,
    )
    row += 1

    phases = ["P1", "P1", "P2", "P2", "P3", "P3", "P3", "P4", "P4", "P4", "P5", "P5"]

    # 売上 (10社優遇 + 新規保守的試算)
    revenue_rows = [
        ("【売上】見学会予約", "", [0, 50000, 125000, 175000, 225000, 275000, 325000, 375000, 425000, 475000, 500000, 550000]),
        ("【売上】成果報酬", "", [0, 125000, 125000, 125000, 250000, 250000, 375000, 500000, 500000, 625000, 750000, 750000]),
        ("【売上】撮影プラン", "", [0, 0, 75000, 150000, 225000, 225000, 300000, 300000, 375000, 375000, 450000, 450000]),
        ("【売上】月額掲載", "", [0, 0, 0, 0, 30000, 90000, 180000, 240000, 270000, 300000, 360000, 390000]),
        ("【売上】SaaS", "", [0, 0, 0, 0, 0, 0, 0, 75000, 150000, 225000, 300000, 375000]),
    ]
    for label, _, vals in revenue_rows:
        R(ws, row, [label, ""] + [f"¥{v:,}" for v in vals], fill=CREAM, align=RIGHT)
        ws.cell(row=row, column=1).alignment = LEFT
        row += 1

    # 売上合計
    totals = [sum(vals[i] for _, _, vals in revenue_rows) for i in range(12)]
    R(
        ws, row,
        ["売上合計", ""] + [f"¥{t:,}" for t in totals],
        fill=ORANGE, align=RIGHT, bold=True,
    )
    ws.cell(row=row, column=1).alignment = LEFT
    ws.cell(row=row, column=1).font = WHITE_BOLD
    for col in range(3, 15):
        ws.cell(row=row, column=col).font = WHITE_BOLD
    revenue_total_row = row
    row += 2

    # 変動原価
    cost_rows = [
        ("【原価】見学会予約 (10%)", "", [0, 5000, 12500, 17500, 22500, 27500, 32500, 37500, 42500, 47500, 50000, 55000]),
        ("【原価】成果報酬 (2%)", "", [0, 2500, 2500, 2500, 5000, 5000, 7500, 10000, 10000, 12500, 15000, 15000]),
        ("【原価】撮影プラン (53%)", "", [0, 0, 40000, 80000, 120000, 120000, 160000, 160000, 200000, 200000, 240000, 240000]),
        ("【原価】月額掲載 (10%)", "", [0, 0, 0, 0, 3000, 9000, 18000, 24000, 27000, 30000, 36000, 39000]),
        ("【原価】SaaS (16%)", "", [0, 0, 0, 0, 0, 0, 0, 12000, 24000, 36000, 48000, 60000]),
    ]
    for label, _, vals in cost_rows:
        R(ws, row, [label, ""] + [f"¥{v:,}" for v in vals], align=RIGHT)
        ws.cell(row=row, column=1).alignment = LEFT
        row += 1
    cost_totals = [sum(vals[i] for _, _, vals in cost_rows) for i in range(12)]
    R(
        ws, row,
        ["変動原価 合計", ""] + [f"¥{t:,}" for t in cost_totals],
        fill=YELLOW, align=RIGHT, bold=True,
    )
    ws.cell(row=row, column=1).alignment = LEFT
    row += 2

    # 粗利
    gross = [totals[i] - cost_totals[i] for i in range(12)]
    R(
        ws, row,
        ["粗利", ""] + [f"¥{g:,}" for g in gross],
        fill=GREEN, align=RIGHT, bold=True,
    )
    ws.cell(row=row, column=1).alignment = LEFT
    row += 1
    R(
        ws, row,
        ["粗利率", ""] + [f"{(g/t*100):.0f}%" if t else "−" for g, t in zip(gross, totals)],
        fill=GREEN, align=RIGHT,
    )
    ws.cell(row=row, column=1).alignment = LEFT
    row += 2

    # 固定費想定
    fixed_monthly = 600_000
    R(
        ws, row,
        ["固定費 (人件費・家賃・SaaS)", ""] + [f"¥{fixed_monthly:,}"] * 12,
        fill=RED, align=RIGHT,
    )
    ws.cell(row=row, column=1).alignment = LEFT
    row += 1

    # 営業利益
    op = [gross[i] - fixed_monthly for i in range(12)]
    R(
        ws, row,
        ["営業利益"] + [""] + [f"¥{o:,}" for o in op],
        fill=BLUE, align=RIGHT, bold=True,
    )
    ws.cell(row=row, column=1).alignment = LEFT
    row += 2

    # Phase row
    phase_row = 3
    for i, p in enumerate(phases, start=3):
        ws.cell(row=phase_row, column=i, value=p).fill = CREAM

    ws.freeze_panes = "C3"

    # ── ④ 原価・変動費 ──
    ws = wb.create_sheet("④ 原価・変動費")
    row = T(ws, 1, "④ プラン別 原価・粗利", 5, "プラットフォームの採算性評価")
    H(ws, row, ["プラン", "正規価格", "変動原価", "粗利", "粗利率"], [32, 18, 18, 18, 12])
    row += 1
    cost_detail = [
        ("見学会予約 (1件)", 50_000, 5_000, 45_000, "90%"),
        ("見学会予約 (10社半額)", 25_000, 5_000, 20_000, "80%"),
        ("成果報酬 (¥3000万契約想定)", 900_000, 20_000, 880_000, "98%"),
        ("撮影プラン (1本)", 150_000, 80_000, 70_000, "47%"),
        ("撮影プラン (10社半額)", 75_000, 80_000, -5_000, "−7%"),
        ("月額ベーシック", 30_000, 3_000, 27_000, "90%"),
        ("月額ベーシック (10社半額)", 15_000, 3_000, 12_000, "80%"),
        ("月額プロ", 80_000, 10_000, 70_000, "88%"),
        ("SaaS ライト", 50_000, 8_000, 42_000, "84%"),
        ("SaaS ライト (10社半額)", 25_000, 8_000, 17_000, "68%"),
        ("SaaS プロ", 150_000, 30_000, 120_000, "80%"),
    ]
    for name, price, var, profit, rate in cost_detail:
        fill = RED if profit <= 0 else None
        R(
            ws, row,
            [name, f"¥{price:,}", f"¥{var:,}", f"¥{profit:,}", rate],
            align=CENTER, fill=fill,
        )
        row += 1
    row += 1
    note = ws.cell(row=row, column=1, value="※ 10社優遇の撮影プラン (¥75,000/本) は原価割れの可能性あり。Phase 2 入ったら撮影コストの削減交渉が必要。")
    note.font = SMALL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.freeze_panes = "A3"

    # ── ⑤ 固定費 年間計画 ──
    ws = wb.create_sheet("⑤ 固定費")
    row = T(ws, 1, "⑤ 固定費 年間計画", 4, "毎月発生する固定費の想定")
    H(ws, row, ["項目", "月額", "年額", "備考"], [30, 16, 18, 40])
    row += 1
    fixed = [
        ("代表・PM 人件費", 300_000, 3_600_000, "代表 + PM 1名"),
        ("エンジニア 人件費", 150_000, 1_800_000, "Phase 2 から業務委託1名想定"),
        ("営業 人件費", 0, 0, "Phase 1 は代表が兼任"),
        ("Vercel Pro", 3_000, 36_000, "Phase 1 は Hobby で OK、Phase 2 から Pro"),
        ("Supabase Pro", 4_000, 48_000, "Phase 2 以降"),
        ("Sentry", 3_000, 36_000, "Phase 1 から"),
        ("GA4 / Search Console", 0, 0, "無料"),
        ("ドメイン", 1_500, 18_000, "payhome.jp"),
        ("経理ソフト (freee)", 3_000, 36_000, ""),
        ("オフィス家賃", 80_000, 960_000, "鹿児島拠点"),
        ("通信費", 10_000, 120_000, ""),
        ("水道光熱費", 15_000, 180_000, ""),
        ("交通費", 20_000, 240_000, "工務店訪問"),
        ("広告宣伝費", 10_000, 120_000, "Phase 1 は YouTube のオーガニックで"),
    ]
    sum_m = 0
    for name, m, y, note in fixed:
        R(ws, row, [name, f"¥{m:,}", f"¥{y:,}", note])
        sum_m += m
        row += 1
    R(ws, row, ["合計", f"¥{sum_m:,}", f"¥{sum_m*12:,}", ""], fill=ORANGE, bold=True)
    ws.cell(row=row, column=1).font = WHITE_BOLD
    ws.cell(row=row, column=2).font = WHITE_BOLD
    ws.cell(row=row, column=3).font = WHITE_BOLD
    ws.freeze_panes = "A3"

    # ── ⑥ 税務・請求フロー ──
    ws = wb.create_sheet("⑥ 請求フロー")
    row = T(ws, 1, "⑥ 請求書発行から入金確認までの手順", 4, "毎月1日〜月末までのサイクル")
    H(ws, row, ["ステップ", "期日", "担当", "作業内容"], [10, 14, 12, 60])
    row += 1
    flow = [
        ("1", "前月末", "PM", "月次請求台帳 (①シート) に当月発生する請求を記入"),
        ("2", "月初3営業日以内", "経理", "請求書テンプレートで各社の請求書を作成"),
        ("3", "月初3営業日以内", "経理", "請求書を PDF で工務店様へメール送信"),
        ("4", "送信当日", "経理", "入金管理シート (②) に請求行を追加"),
        ("5", "月末", "経理", "入金確認 → 入金管理シートを更新"),
        ("6", "月末翌日", "経理", "未入金案件を PM へエスカレーション"),
        ("7", "月末", "PM", "月次P&L シート (③) を実績値で更新"),
        ("8", "月末翌日", "PM", "代表・経理へ月次レポート共有"),
        ("9", "四半期ごと", "代表・経理", "消費税・法人税の納付準備"),
        ("10", "年度末", "代表・経理", "決算書作成 (外部税理士に依頼)"),
    ]
    for f in flow:
        R(ws, row, list(f), align=LEFT_TOP)
        row += 1
    ws.freeze_panes = "A3"

    out = OUT_KEIRI / "請求・P&L管理.xlsx"
    wb.save(out)
    return out


# ══════════════════════════════════════════════════════════════════
# 2. 動線設計シート (フェーズ別 KPI 逆算)
# ══════════════════════════════════════════════════════════════════

def build_funnel_design() -> Path:
    wb = Workbook()

    # ── 目次 ──
    ws = wb.active
    ws.title = "目次"
    row = T(
        ws, 1, "ぺいほーむ 動線設計シート", 3,
        "目標収益 (件数) から逆算した動線設計とKPI管理。Phase 毎にどの動線をどこまで伸ばすかを定義",
    )
    H(ws, row, ["#", "シート名", "用途"], [6, 32, 60])
    row += 1
    toc = [
        ("①", "収益→件数 逆算", "目標売上からフェーズ毎の必要件数を逆算"),
        ("②", "動線マップ 全体像", "ユーザーの入口から収益発生までの動線を可視化"),
        ("③", "動線別 Phase1 KPI", "見学会予約動線 / AI診断動線 / 動画動線 / 相談動線"),
        ("④", "動線別 Phase2 KPI", "P1の継続動線 + 記事動線 + 撮影プラン動線"),
        ("⑤", "動線別 Phase3 KPI", "P2の継続動線 + 物件動線 + 月額掲載動線"),
        ("⑥", "動線別 Phase4 KPI", "P3の継続動線 + AI質問動線 + SaaS動線"),
        ("⑦", "動線別 Phase5 KPI", "全国展開期の動線"),
        ("⑧", "ファネル逆算 (詳細)", "UU→診断→予約→成約の各ステップ転換率"),
        ("⑨", "動線施策 優先度", "どの動線を優先してテコ入れするかのマトリクス"),
    ]
    for t in toc:
        R(ws, row, list(t))
        row += 1

    # ── ① 収益→件数 逆算 ──
    ws = wb.create_sheet("① 収益→件数 逆算")
    row = T(
        ws, 1, "① 収益目標から件数への逆算", 8,
        "Phase別の月次売上目標と、それを達成するために必要な件数",
    )
    H(
        ws, row,
        ["Phase", "月", "売上目標", "見学会予約 件数", "成果報酬 件数", "撮影 本数", "月額掲載 社数", "SaaS 社数"],
        [6, 8, 14, 14, 14, 12, 14, 12],
    )
    row += 1

    # (phase, month, target_rev, booking_cnt, success_cnt, filming_cnt, sub_cnt, saas_cnt)
    targets = [
        ("P1", "5月", 0, 0, 0, 0, 0, 0),       # 全て無料の優遇期間
        ("P1", "6月", 175_000, 1, 1, 0, 0, 0),
        ("P2", "7月", 325_000, 3, 1, 1, 0, 0),
        ("P2", "8月", 450_000, 4, 1, 2, 0, 0),
        ("P3", "9月", 730_000, 5, 2, 3, 1, 0),
        ("P3", "10月", 840_000, 6, 2, 3, 3, 0),
        ("P3", "11月", 1_180_000, 7, 3, 4, 6, 0),
        ("P4", "12月", 1_490_000, 8, 4, 4, 8, 1),
        ("P4", "1月", 1_720_000, 9, 4, 5, 9, 2),
        ("P4", "2月", 2_000_000, 10, 5, 5, 10, 3),
        ("P5", "3月", 2_360_000, 11, 6, 6, 12, 4),
        ("P5", "4月", 2_515_000, 12, 6, 6, 13, 5),
    ]
    for p, m, rev, b, s, f, sub, saas in targets:
        fill = GRAY if rev == 0 else (CREAM if rev < 1_000_000 else GREEN)
        R(
            ws, row,
            [p, m, f"¥{rev:,}", b, s, f, sub, saas],
            align=CENTER, fill=fill,
        )
        row += 1

    # Totals
    t_rev = sum(t[2] for t in targets)
    R(
        ws, row,
        ["年間", "", f"¥{t_rev:,}", sum(t[3] for t in targets),
         sum(t[4] for t in targets), sum(t[5] for t in targets),
         sum(t[6] for t in targets), sum(t[7] for t in targets)],
        align=CENTER, fill=ORANGE, bold=True,
    )
    for c in range(1, 9):
        ws.cell(row=row, column=c).font = WHITE_BOLD
    ws.freeze_panes = "A3"

    # ── ② 動線マップ 全体像 ──
    ws = wb.create_sheet("② 動線マップ")
    row = T(
        ws, 1, "② 動線マップ (入口→収益)", 5,
        "ユーザーがどこから入って、どのページを経由して、どこで収益が発生するか",
    )
    H(ws, row, ["入口", "第1ステップ", "第2ステップ", "ゴール", "収益発生"], [20, 26, 26, 22, 22])
    row += 1
    routes = [
        ("YouTube (チャンネル)", "動画説明欄リンク", "ぺいほーむTOP or 工務店ページ", "見学会予約", "見学会予約費 + 成果報酬"),
        ("Google検索 (指名)", "ぺいほーむTOP", "工務店一覧 → 詳細", "見学会予約", "見学会予約費 + 成果報酬"),
        ("Google検索 (平屋 鹿児島)", "工務店一覧 or 事例", "工務店詳細 → 見学会", "見学会予約", "見学会予約費 + 成果報酬"),
        ("SNS (X/Insta/Threads)", "ぺいほーむTOP", "AI診断", "診断結果 → 予約", "見学会予約費 + 成果報酬"),
        ("AI家づくり診断 (広告)", "診断フォーム", "結果ページ", "お気に入り or 予約", "見学会予約費 + 成果報酬"),
        ("LINE公式", "LINE内リンク", "ぺいほーむTOP → AI診断", "相談予約", "送客先の工務店からの成果報酬"),
        ("無料相談 (TOPCTA)", "相談フォーム", "スタッフ対応", "見学会予約 or 契約", "見学会予約費 + 成果報酬"),
        ("デジタルカタログDL", "カタログページ", "会員登録", "メルマガ → 診断・予約", "長期的送客"),
        ("Phase2: 業界ニュース", "ニュース記事", "工務店詳細", "見学会予約", "撮影プラン販売機会"),
        ("Phase3: 建売物件", "物件詳細", "工務店詳細", "物件問合せ", "月額掲載料"),
        ("Phase4: 匿名AI質問", "質問投稿", "工務店回答", "リード → 予約", "SaaSプラン"),
    ]
    for r in routes:
        R(ws, row, list(r))
        row += 1
    ws.freeze_panes = "A3"

    # ── ③-⑦ 動線別 Phase別 KPI ──
    def build_phase_sheet(
        wb: Workbook, sheet_name: str, title: str, subtitle: str, routes_data: list
    ):
        ws = wb.create_sheet(sheet_name)
        row = T(ws, 1, title, 9, subtitle)
        H(
            ws, row,
            [
                "動線名",
                "入口",
                "主要KPI",
                "月間目標 (人数)",
                "転換率",
                "月次リード",
                "見学会予約",
                "収益インパクト",
                "担当",
            ],
            [20, 22, 22, 14, 12, 14, 14, 20, 12],
        )
        row += 1
        for r in routes_data:
            R(ws, row, list(r), align=LEFT_TOP)
            row += 1
        ws.freeze_panes = "A3"

    # Phase 1 KPI
    p1_routes = [
        (
            "YouTube動線",
            "YouTubeチャンネル",
            "動画視聴数 → ぺいほーむ遷移率",
            "3,000 UU",
            "5%",
            "150",
            "3件",
            "¥75,000〜¥125,000/月",
            "代表",
        ),
        (
            "AI診断動線",
            "TOPのAI診断CTA",
            "診断完了数 → 工務店詳細遷移率",
            "200 UU",
            "30%",
            "60",
            "2件",
            "¥50,000〜¥100,000/月",
            "PM",
        ),
        (
            "工務店検索動線",
            "/builders 一覧",
            "工務店詳細閲覧数",
            "500 UU",
            "8%",
            "40",
            "1件",
            "¥25,000〜¥50,000/月",
            "営業",
        ),
        (
            "見学会動線",
            "/event 一覧",
            "見学会予約フォーム完了率",
            "200 UU",
            "15%",
            "30",
            "3件",
            "¥75,000〜¥125,000/月",
            "営業",
        ),
        (
            "無料相談動線",
            "TOP相談CTA",
            "相談フォーム送信数",
            "100 UU",
            "5%",
            "5",
            "1件",
            "¥25,000〜¥50,000/月",
            "営業",
        ),
        (
            "事例ライブラリ動線",
            "/case-studies",
            "事例閲覧 → 工務店遷移率",
            "300 UU",
            "10%",
            "30",
            "0件",
            "間接送客のみ",
            "PM",
        ),
    ]
    build_phase_sheet(
        wb, "③ Phase1 動線KPI",
        "③ Phase 1 動線別KPI (2026/05/01〜06/30)",
        "MVP 17画面の6動線。目標月間予約10件を各動線で分担",
        p1_routes,
    )

    # Phase 2 KPI (adds 記事動線 + 撮影プラン動線)
    p2_routes = p1_routes + [
        (
            "業界ニュース動線",
            "/news 一覧",
            "ニュース閲覧 → 工務店遷移率",
            "800 UU",
            "10%",
            "80",
            "2件",
            "¥50,000〜¥100,000/月",
            "PM",
        ),
        (
            "記事動線",
            "/articles 一覧",
            "記事閲覧 → 工務店遷移率",
            "1,200 UU",
            "8%",
            "96",
            "3件",
            "¥75,000〜¥150,000/月",
            "PM",
        ),
        (
            "撮影プラン動線",
            "工務店向けアウトバウンド",
            "撮影プラン申込数",
            "10社アプローチ",
            "20%",
            "2社",
            "−",
            "¥150,000〜¥300,000/月",
            "代表",
        ),
    ]
    build_phase_sheet(
        wb, "④ Phase2 動線KPI",
        "④ Phase 2 動線別KPI (2026/07/01〜08/31)",
        "記事・ニュース動線を追加。撮影プラン販売で客単価UP",
        p2_routes,
    )

    # Phase 3 KPI (adds 物件動線 + 月額掲載動線)
    p3_routes = p2_routes + [
        (
            "建売物件動線",
            "/sale-homes 一覧",
            "物件詳細 → 問合せ率",
            "1,500 UU",
            "6%",
            "90",
            "4件",
            "¥100,000〜¥200,000/月",
            "PM",
        ),
        (
            "土地情報動線",
            "/lands 一覧",
            "土地詳細 → 問合せ率",
            "800 UU",
            "5%",
            "40",
            "2件",
            "¥50,000〜¥100,000/月",
            "PM",
        ),
        (
            "特集動線",
            "/features 一覧",
            "特集閲覧 → 工務店遷移率",
            "1,000 UU",
            "10%",
            "100",
            "3件",
            "¥75,000〜¥150,000/月",
            "PM",
        ),
        (
            "月額掲載動線",
            "工務店向けアウトバウンド",
            "月額プラン加入社数",
            "15社アプローチ",
            "30%",
            "5社",
            "−",
            "¥150,000〜¥300,000/月",
            "営業",
        ),
    ]
    build_phase_sheet(
        wb, "⑤ Phase3 動線KPI",
        "⑤ Phase 3 動線別KPI (2026/09/01〜11/30)",
        "物件情報3動線を追加。月額掲載販売でストック収益開始",
        p3_routes,
    )

    # Phase 4 KPI (adds AI質問動線 + SaaS動線)
    p4_routes = p3_routes + [
        (
            "AI質問動線",
            "/mypage/questions",
            "質問投稿数 → 回答率",
            "200 質問",
            "80%",
            "160",
            "4件",
            "¥100,000〜¥200,000/月",
            "PM",
        ),
        (
            "ウェビナー動線",
            "/webinar 一覧",
            "参加申込数 → 成約率",
            "500 UU",
            "8%",
            "40",
            "3件",
            "¥75,000〜¥150,000/月",
            "PM",
        ),
        (
            "SaaS動線",
            "工務店向けアウトバウンド",
            "SaaSプラン加入社数",
            "20社アプローチ",
            "25%",
            "5社",
            "−",
            "¥250,000〜¥500,000/月",
            "営業",
        ),
    ]
    build_phase_sheet(
        wb, "⑥ Phase4 動線KPI",
        "⑥ Phase 4 動線別KPI (2026/12/01〜2027/02/28)",
        "AI質問・ウェビナー動線を追加。SaaS販売で MRR 急成長期",
        p4_routes,
    )

    # Phase 5 KPI (全国展開)
    p5_routes = [
        (
            "県別SEO動線 (福岡)",
            "/area/fukuoka",
            "検索流入 → 工務店遷移率",
            "3,000 UU",
            "10%",
            "300",
            "15件",
            "¥375,000〜¥750,000/月",
            "PM",
        ),
        (
            "県別SEO動線 (熊本・宮崎・大分・佐賀)",
            "/area/[prefecture]",
            "検索流入 → 工務店遷移率",
            "5,000 UU",
            "8%",
            "400",
            "20件",
            "¥500,000〜¥1,000,000/月",
            "PM",
        ),
        (
            "既存動線 (全動線合算)",
            "Phase1-4の全動線",
            "継続運用",
            "40,000 UU",
            "−",
            "−",
            "50件",
            "¥2,000,000〜¥3,000,000/月",
            "全員",
        ),
        (
            "有料広告動線",
            "Google Ads / Meta Ads",
            "広告CPA",
            "5,000 UU",
            "5%",
            "250",
            "10件",
            "¥250,000〜¥500,000/月",
            "代表",
        ),
    ]
    build_phase_sheet(
        wb, "⑦ Phase5 動線KPI",
        "⑦ Phase 5 動線別KPI (2027/03〜)",
        "全国展開。県別SEO + 有料広告で総UUを4万→8万へ",
        p5_routes,
    )

    # ── ⑧ ファネル逆算 (詳細) ──
    ws = wb.create_sheet("⑧ ファネル逆算")
    row = T(
        ws, 1, "⑧ ファネル逆算 (詳細)", 6,
        "UU → AI診断 → 工務店詳細 → 見学会予約 → 成約 の各ステップの転換率と必要数",
    )
    H(
        ws, row,
        ["Phase", "目標予約数/月", "必要 成約件数", "必要 予約数", "必要 AI診断数", "必要 UU"],
        [6, 18, 16, 16, 16, 16],
    )
    row += 1
    # 転換率仮定:
    # UU → AI診断: 4% (P1で200/5000)
    # AI診断 → 工務店詳細: 50%
    # 工務店詳細 → 見学会予約: 10%
    # 見学会予約 → 成約: 30%
    funnel = [
        ("P1", 10, 3, 10, 250, 6_250),
        ("P2", 20, 6, 20, 500, 12_500),
        ("P3", 40, 12, 40, 1_000, 25_000),
        ("P4", 55, 17, 55, 1_375, 34_375),
        ("P5", 70, 21, 70, 1_750, 43_750),
    ]
    for phase, b, s, req_b, req_diag, req_uu in funnel:
        R(
            ws, row,
            [phase, f"{b}件", f"{s}件", f"{req_b}件", f"{req_diag:,}", f"{req_uu:,}"],
            align=CENTER, fill=CREAM,
        )
        row += 1

    row += 1
    note_row = row
    c = ws.cell(row=note_row, column=1, value="前提転換率")
    c.font = BOLD
    row += 1
    conversions = [
        ("UU → AI診断", "4%", "TOP CTA からの誘導"),
        ("AI診断 → 工務店詳細", "50%", "診断結果の工務店カードから"),
        ("工務店詳細 → 見学会予約", "10%", "見学会予約CTAの完了率"),
        ("見学会予約 → 成約", "30%", "予約後の契約率 (工務店依存)"),
    ]
    H(ws, row, ["ステップ", "転換率", "備考"], [32, 14, 50])
    row += 1
    for s in conversions:
        R(ws, row, list(s))
        row += 1
    ws.freeze_panes = "A3"

    # ── ⑨ 動線施策 優先度 ──
    ws = wb.create_sheet("⑨ 施策優先度")
    row = T(
        ws, 1, "⑨ 動線施策 優先度マトリクス", 5,
        "収益インパクト x 実装難易度 で動線施策を優先度付け",
    )
    H(
        ws, row,
        ["優先度", "施策", "動線", "収益インパクト", "実装難易度"],
        [10, 40, 20, 18, 14],
    )
    row += 1
    actions = [
        ("S", "YouTube動画の冒頭3秒でぺいほーむ紹介を毎本挿入", "YouTube動線", "高", "低"),
        ("S", "見学会予約CTAをモバイルで追従フッター固定", "見学会動線", "高", "低"),
        ("S", "AI診断完了後の工務店カードに「見学会予約」ボタン併設", "AI診断動線", "高", "低"),
        ("A", "動画カードにサムネオーバーレイで「この家が気になる」CTA", "YouTube動線", "高", "中"),
        ("A", "TOP の1st view に「見学会予約」CTA併設", "見学会動線", "高", "中"),
        ("A", "AI診断の質問数を10→7問に減らす (離脱率改善)", "AI診断動線", "中", "中"),
        ("A", "工務店詳細ページに「この工務店の見学会カレンダー」埋込", "工務店検索動線", "高", "中"),
        ("B", "無料相談のフォームを1画面でLINE連携", "無料相談動線", "中", "中"),
        ("B", "事例カードの下に「同じ工務店の動画を見る」リンク", "事例動線", "中", "低"),
        ("B", "Phase2: ニュース記事内のCTA実装 (工務店回遊)", "記事動線", "中", "中"),
        ("B", "Phase2: 撮影プラン申込フォームの導線整備", "撮影動線", "高", "中"),
        ("C", "Phase3: 物件詳細の「近くの工務店」自動レコメンド", "物件動線", "中", "高"),
        ("C", "Phase3: 月額プランの体験版 (1ヶ月無料)", "月額動線", "中", "低"),
        ("C", "Phase4: AI質問のリアルタイム通知Push", "AI質問動線", "中", "高"),
        ("C", "Phase4: SaaSデモ環境のセルフサインアップ", "SaaS動線", "高", "高"),
        ("D", "Phase5: 県別ランディングページ x 5県", "県別SEO", "高", "高"),
        ("D", "Phase5: Google広告のリマーケティング", "有料広告", "中", "中"),
    ]
    fill_map = {"S": RED, "A": YELLOW, "B": CREAM, "C": GRAY, "D": BLUE}
    for priority, action, route, impact, diff in actions:
        R(
            ws, row,
            [priority, action, route, impact, diff],
            fill=fill_map.get(priority),
        )
        row += 1
    ws.freeze_panes = "A3"

    out = OUT_MGMT / "動線設計シート.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    p1 = build_accounting()
    p2 = build_funnel_design()
    print("Created:", p1)
    print("Created:", p2)
